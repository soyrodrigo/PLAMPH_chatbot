#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 DIAGNÓSTICO DEL EXCEL - VERIFICAR QUÉ DATOS ESTÁ LEYENDO
===========================================================

Este script te ayudará a identificar por qué la gráfica no muestra
los datos reales de tu archivo Excel.
"""

import os
import sys
from datetime import datetime
import openpyxl

def verificar_archivos_excel():
    """Verifica qué archivos Excel existen y dónde están"""
    print("🔍 === VERIFICACIÓN DE ARCHIVOS EXCEL ===")
    print(f"📁 Directorio actual: {os.getcwd()}")
    
    # Posibles ubicaciones del archivo
    ubicaciones_posibles = [
        "inventario_materiales.xlsx",
        "datos/inventario_materiales.xlsx", 
        "./datos/inventario_materiales.xlsx",
        os.path.join("datos", "inventario_materiales.xlsx")
    ]
    
    archivos_encontrados = []
    
    for ubicacion in ubicaciones_posibles:
        if os.path.exists(ubicacion):
            tamaño = os.path.getsize(ubicacion) / 1024  # KB
            fecha_mod = datetime.fromtimestamp(os.path.getmtime(ubicacion))
            archivos_encontrados.append({
                'ruta': ubicacion,
                'tamaño': tamaño,
                'fecha_modificacion': fecha_mod
            })
            print(f"✅ ENCONTRADO: {ubicacion}")
            print(f"   📏 Tamaño: {tamaño:.1f} KB")
            print(f"   📅 Última modificación: {fecha_mod.strftime('%d/%m/%Y %H:%M:%S')}")
        else:
            print(f"❌ NO EXISTE: {ubicacion}")
    
    return archivos_encontrados

def leer_contenido_excel(ruta_archivo):
    """Lee y muestra el contenido real del archivo Excel"""
    print(f"\n📖 === LEYENDO CONTENIDO DE: {ruta_archivo} ===")
    
    try:
        libro = openpyxl.load_workbook(ruta_archivo)
        hoja = libro.active
        
        print(f"📋 Nombre de la hoja: {hoja.title}")
        print(f"📊 Filas con datos: {hoja.max_row}")
        print(f"📊 Columnas con datos: {hoja.max_column}")
        
        # Mostrar encabezados
        print(f"\n📝 ENCABEZADOS (Fila 4):")
        for col in range(1, hoja.max_column + 1):
            valor = hoja.cell(row=4, column=col).value
            print(f"   Columna {col}: {valor}")
        
        # Mostrar primeras 10 filas de datos
        print(f"\n📋 PRIMERAS 10 FILAS DE DATOS:")
        for row in range(5, min(hoja.max_row + 1, 15)):  # Filas 5-14
            fila_datos = []
            for col in range(1, min(hoja.max_column + 1, 8)):  # Primeras 7 columnas
                valor = hoja.cell(row=row, column=col).value
                fila_datos.append(str(valor) if valor else "")
            
            if any(fila_datos):  # Solo mostrar si hay datos
                print(f"   Fila {row}: {' | '.join(fila_datos)}")
        
        # Análisis específico de materiales
        print(f"\n🔍 ANÁLISIS DE MATERIALES:")
        materiales_encontrados = {}
        
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value  # Columna C (Material)
            movimiento = hoja.cell(row=row, column=5).value  # Columna E (Movimiento)
            cantidad = hoja.cell(row=row, column=6).value  # Columna F (Cantidad)
            
            if material and movimiento and cantidad:
                print(f"   📦 {material} | {movimiento} | {cantidad}")
                
                # Calcular stock como lo hace el sistema
                if material not in materiales_encontrados:
                    materiales_encontrados[material] = 0
                
                try:
                    cantidad_num = float(str(cantidad).replace(",", "."))
                    if "Entrada" in str(movimiento):
                        materiales_encontrados[material] += cantidad_num
                    elif "Salida" in str(movimiento):
                        materiales_encontrados[material] -= cantidad_num
                except:
                    print(f"      ⚠️ Error procesando cantidad: {cantidad}")
        
        print(f"\n📊 STOCK CALCULADO:")
        for material, stock in materiales_encontrados.items():
            print(f"   📦 {material}: {stock:.1f}")
        
        return materiales_encontrados
        
    except Exception as e:
        print(f"❌ Error leyendo archivo: {e}")
        return None

def verificar_modulo_excel_manager():
    """Verifica qué está leyendo el módulo ExcelManager"""
    print(f"\n🔧 === VERIFICANDO MÓDULO EXCEL_MANAGER ===")
    
    try:
        # Intentar importar
        sys.path.insert(0, '.')
        from modules.excel_manager import ExcelManager
        print("✅ ExcelManager importado correctamente")
        
        # Verificar qué archivo está usando
        try:
            # Intentar obtener la ruta del archivo que usa
            import modules.config as config
            archivo_materiales = getattr(config, 'ARCHIVO_EXCEL_MATERIALES', 'No definido')
            print(f"📁 Archivo configurado: {archivo_materiales}")
        except:
            print("⚠️ No se pudo obtener configuración del archivo")
        
        # Probar funciones del ExcelManager
        print(f"\n🧪 PROBANDO FUNCIONES DEL EXCELMANAGER:")
        
        # Verificar archivos
        ExcelManager.verificar_y_crear_archivos()
        print("✅ verificar_y_crear_archivos() ejecutado")
        
        # Obtener stock
        stock = ExcelManager.obtener_stock_materiales()
        if stock:
            print(f"📊 Stock obtenido por ExcelManager:")
            for material, cantidad in stock.items():
                print(f"   📦 {material}: {cantidad:.1f}")
        else:
            print("❌ ExcelManager no devolvió stock")
        
        # Obtener últimos movimientos
        movimientos = ExcelManager.obtener_ultimos_movimientos(5)
        if movimientos:
            print(f"\n📋 Últimos movimientos según ExcelManager:")
            for mov in movimientos:
                print(f"   {mov}")
        else:
            print("❌ No hay movimientos según ExcelManager")
        
        return stock
        
    except Exception as e:
        print(f"❌ Error con ExcelManager: {e}")
        import traceback
        traceback.print_exc()
        return None

def comparar_datos(datos_directos, datos_excel_manager):
    """Compara los datos leídos directamente vs ExcelManager"""
    print(f"\n⚖️ === COMPARACIÓN DE DATOS ===")
    
    if not datos_directos and not datos_excel_manager:
        print("❌ No hay datos para comparar")
        return
    
    if not datos_directos:
        print("❌ No hay datos leídos directamente del archivo")
        datos_directos = {}
        
    if not datos_excel_manager:
        print("❌ No hay datos del ExcelManager")
        datos_excel_manager = {}
    
    print(f"\n📋 COMPARACIÓN:")
    print(f"{'Material':<15} {'Directo':<10} {'ExcelManager':<15} {'¿Coincide?'}")
    print("-" * 55)
    
    todos_materiales = set(list(datos_directos.keys()) + list(datos_excel_manager.keys()))
    
    coincidencias = 0
    total = len(todos_materiales)
    
    for material in todos_materiales:
        directo = datos_directos.get(material, 0)
        excel_mgr = datos_excel_manager.get(material, 0)
        coincide = "✅" if abs(directo - excel_mgr) < 0.01 else "❌"
        
        if coincide == "✅":
            coincidencias += 1
        
        print(f"{material:<15} {directo:<10.1f} {excel_mgr:<15.1f} {coincide}")
    
    if total > 0:
        porcentaje = (coincidencias / total) * 100
        print(f"\n📊 RESULTADO: {coincidencias}/{total} coincidencias ({porcentaje:.1f}%)")
        
        if porcentaje < 100:
            print("🔥 ¡PROBLEMA IDENTIFICADO! Los datos no coinciden")
            print("💡 El ExcelManager está leyendo datos diferentes al archivo real")
        else:
            print("✅ Los datos coinciden perfectamente")
    else:
        print("⚠️ No hay materiales para comparar")

def main():
    """Función principal del diagnóstico"""
    print("🔍 DIAGNÓSTICO COMPLETO DEL SISTEMA EXCEL")
    print("=" * 60)
    print(f"⏰ Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # 1. Verificar archivos
    archivos = verificar_archivos_excel()
    
    if not archivos:
        print("\n❌ NO SE ENCONTRÓ NINGÚN ARCHIVO EXCEL")
        print("💡 Esto explica por qué la gráfica muestra datos incorrectos")
        return
    
    # 2. Leer archivo más reciente
    archivo_principal = max(archivos, key=lambda x: x['fecha_modificacion'])
    print(f"\n📖 Analizando archivo más reciente: {archivo_principal['ruta']}")
    
    datos_directos = leer_contenido_excel(archivo_principal['ruta'])
    
    # 3. Verificar ExcelManager  
    datos_excel_manager = verificar_modulo_excel_manager()
    
    # 4. Comparar datos
    comparar_datos(datos_directos, datos_excel_manager)
    
    # 5. Recomendaciones
    print(f"\n💡 === RECOMENDACIONES ===")
    
    if not datos_directos:
        print("1. ❌ Tu archivo Excel está vacío o corrupto")
        print("   Agrega algunos datos manualmente o usa el bot")
    elif not datos_excel_manager:
        print("2. ❌ ExcelManager no funciona correctamente")
        print("   Revisa la configuración del módulo")
    elif datos_directos != datos_excel_manager:
        print("3. ❌ ExcelManager lee datos diferentes")
        print("   Revisa la ruta del archivo en config.py")
        print("   O hay datos de ejemplo interferiendo")
    else:
        print("✅ Todo funciona correctamente")
        print("   El problema puede estar en GraphicsGenerator")
    
    print(f"\n📋 ARCHIVO ANALIZADO: {archivo_principal['ruta']}")
    print(f"📊 REGISTROS ENCONTRADOS: {len(datos_directos) if datos_directos else 0}")
    print(f"⏰ Diagnóstico completado: {datetime.now().strftime('%H:%M:%S')}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠️ Diagnóstico interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error en diagnóstico: {e}")
        import traceback
        traceback.print_exc()
