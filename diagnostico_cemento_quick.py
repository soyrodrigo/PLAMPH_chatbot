#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 DIAGNÓSTICO RÁPIDO - PROBLEMA GRÁFICA CEMENTO
==============================================
Ejecuta esto para verificar el estado actual del problema
"""

import os
import sys
from datetime import datetime

def diagnosticar_problema():
    """Diagnostica el problema con la gráfica de cemento"""
    
    print("🔍 === DIAGNÓSTICO GRÁFICA CEMENTO ===")
    print(f"⏰ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 50)
    
    # 1. Verificar estructura de archivos
    print("\n1️⃣ VERIFICANDO ESTRUCTURA DE ARCHIVOS:")
    
    archivos_criticos = [
        ("modules/", "📁 Carpeta de módulos"),
        ("modules/graphics_generator.py", "📈 Generador de gráficas"),
        ("modules/excel_manager.py", "📊 Gestor Excel"),
        ("datos/inventario_materiales.xlsx", "📄 Archivo de materiales")
    ]
    
    for archivo, descripcion in archivos_criticos:
        if os.path.exists(archivo):
            print(f"   ✅ {descripcion}")
        else:
            print(f"   ❌ {descripcion} - NO ENCONTRADO")
    
    # 2. Probar importación de módulos
    print("\n2️⃣ PROBANDO IMPORTACIÓN DE MÓDULOS:")
    
    try:
        sys.path.append('.')
        from modules.graphics_generator import GraphicsGenerator
        print("   ✅ GraphicsGenerator importado correctamente")
        
        # 3. Probar generación de gráfica
        print("\n3️⃣ PROBANDO GENERACIÓN DE GRÁFICA:")
        
        resultado = GraphicsGenerator.generar_grafica_cemento()
        
        if resultado:
            print(f"   ✅ GRÁFICA GENERADA: {resultado}")
            print("   🎉 ¡EL PROBLEMA ESTÁ RESUELTO!")
            
            if os.path.exists(resultado):
                tamaño = os.path.getsize(resultado) / 1024
                print(f"   📏 Tamaño: {tamaño:.1f} KB")
        else:
            print("   ❌ PROBLEMA CONFIRMADO: No se genera gráfica")
            print("   💡 Necesitas aplicar la corrección")
            
    except ImportError as e:
        print(f"   ❌ Error importando: {e}")
        print("   💡 Verifica que tengas la carpeta modules/")
        
    except Exception as e:
        print(f"   ❌ Error inesperado: {e}")
    
    # 4. Verificar datos de cemento
    print("\n4️⃣ VERIFICANDO DATOS DE CEMENTO EN EXCEL:")
    
    try:
        import openpyxl
        archivo_excel = "datos/inventario_materiales.xlsx"
        
        if os.path.exists(archivo_excel):
            libro = openpyxl.load_workbook(archivo_excel)
            hoja = libro.active
            
            print(f"   📊 Archivo tiene {hoja.max_row} filas")
            
            # Buscar menciones de cemento
            menciones_cemento = 0
            salidas_cemento = 0
            
            for row in range(1, hoja.max_row + 1):
                fila_texto = ""
                for col in range(1, hoja.max_column + 1):
                    valor = hoja.cell(row=row, column=col).value
                    if valor:
                        fila_texto += str(valor).lower() + " "
                
                if "cemento" in fila_texto:
                    menciones_cemento += 1
                    if "salida" in fila_texto or "📉" in fila_texto:
                        salidas_cemento += 1
                        print(f"   📈 Fila {row}: Salida de cemento encontrada")
            
            print(f"   📦 Menciones de cemento: {menciones_cemento}")
            print(f"   📉 Salidas de cemento: {salidas_cemento}")
            
            if salidas_cemento == 0:
                print("   💡 CAUSA PROBABLE: No hay salidas de cemento registradas")
                print("   🔧 SOLUCIÓN: Registra algunas salidas de cemento")
            
        else:
            print("   ❌ Archivo Excel no encontrado")
            
    except Exception as e:
        print(f"   ❌ Error leyendo Excel: {e}")
    
    print("\n" + "=" * 50)
    print("🎯 DIAGNÓSTICO COMPLETADO")
    
    return True

if __name__ == "__main__":
    diagnosticar_problema()