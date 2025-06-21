#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🧹 LIMPIAR DATOS FALSOS DEL EXCEL
================================

Este script eliminará todos los datos de ejemplo/prueba que están
interfiriendo con tus datos reales en el archivo Excel.
"""

import os
import openpyxl
from datetime import datetime

def limpiar_datos_falsos():
    """Elimina datos falsos y conserva solo los datos reales del usuario"""
    
    # Ruta del archivo
    archivo_excel = "datos/inventario_materiales.xlsx"
    
    if not os.path.exists(archivo_excel):
        print(f"❌ No se encuentra el archivo: {archivo_excel}")
        
        # Buscar en otras ubicaciones
        otras_ubicaciones = [
            "inventario_materiales.xlsx",
            "./inventario_materiales.xlsx"
        ]
        
        for ubicacion in otras_ubicaciones:
            if os.path.exists(ubicacion):
                archivo_excel = ubicacion
                print(f"✅ Archivo encontrado en: {archivo_excel}")
                break
        else:
            print("❌ No se encontró ningún archivo Excel")
            return False
    
    print(f"🔍 Analizando archivo: {archivo_excel}")
    
    try:
        # Cargar archivo
        libro = openpyxl.load_workbook(archivo_excel)
        hoja = libro.active
        
        print(f"📊 Filas originales: {hoja.max_row}")
        
        # Identificar materiales falsos (datos de ejemplo)
        materiales_falsos = [
            "Material_0", "Material_1", "Material_2", "Material_3", "Material_4",
            "Material_5", "Material_6", "Material_7", "Material_8", "Material_9",
            "Sistema"  # Usuario "Sistema" también es falso
        ]
        
        # Recopilar filas válidas (datos reales del usuario)
        filas_validas = []
        filas_eliminadas = 0
        
        # Conservar encabezados (filas 1-4)
        for row in range(1, 5):
            fila_datos = []
            for col in range(1, hoja.max_column + 1):
                valor = hoja.cell(row=row, column=col).value
                fila_datos.append(valor)
            filas_validas.append(fila_datos)
        
        # Revisar datos (fila 5 en adelante)
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value  # Columna C - Material
            usuario = hoja.cell(row=row, column=4).value   # Columna D - Usuario
            
            # Verificar si la fila es válida (no es dato falso)
            es_fila_valida = True
            
            # Eliminar si el material es falso
            if material and any(falso in str(material) for falso in materiales_falsos):
                es_fila_valida = False
                print(f"❌ Eliminando material falso: {material}")
            
            # Eliminar si el usuario es "Sistema" (datos de ejemplo)
            if usuario and str(usuario).strip() == "Sistema":
                es_fila_valida = False
                print(f"❌ Eliminando registro del usuario 'Sistema': {material}")
            
            # Eliminar filas con datos de prueba en observaciones
            observaciones = hoja.cell(row=row, column=7).value  # Columna G
            if observaciones and any(texto in str(observaciones).lower() for texto in 
                                   ["prueba", "test", "ejemplo", "registro", "datos de prueba"]):
                es_fila_valida = False
                print(f"❌ Eliminando por observaciones de prueba: {material}")
            
            if es_fila_valida:
                # Conservar esta fila
                fila_datos = []
                for col in range(1, hoja.max_column + 1):
                    valor = hoja.cell(row=row, column=col).value
                    fila_datos.append(valor)
                filas_validas.append(fila_datos)
            else:
                filas_eliminadas += 1
        
        print(f"🧹 Filas eliminadas: {filas_eliminadas}")
        print(f"✅ Filas conservadas: {len(filas_validas) - 4}")  # -4 por los encabezados
        
        # Crear nuevo libro con solo datos válidos
        nuevo_libro = openpyxl.Workbook()
        nueva_hoja = nuevo_libro.active
        nueva_hoja.title = "Inventario Materiales"
        
        # Escribir filas válidas
        for row_idx, fila in enumerate(filas_validas, 1):
            for col_idx, valor in enumerate(fila, 1):
                nueva_hoja.cell(row=row_idx, column=col_idx, value=valor)
        
        # Aplicar formato a encabezados (opcional)
        if len(filas_validas) >= 4:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Formato para encabezados
            for col in range(1, len(filas_validas[3]) + 1):  # Fila 4 son los encabezados
                celda = nueva_hoja.cell(row=4, column=col)
                celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                celda.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Guardar archivo limpio
        archivo_backup = archivo_excel.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        libro.save(archivo_backup)
        print(f"💾 Respaldo creado: {archivo_backup}")
        
        # Guardar archivo limpio
        nuevo_libro.save(archivo_excel)
        print(f"✅ Archivo limpio guardado: {archivo_excel}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error procesando archivo: {e}")
        return False

def verificar_resultado():
    """Verifica que la limpieza fue exitosa"""
    print(f"\n🔍 === VERIFICACIÓN DEL RESULTADO ===")
    
    archivo_excel = "datos/inventario_materiales.xlsx"
    if not os.path.exists(archivo_excel):
        archivo_excel = "inventario_materiales.xlsx"
    
    if not os.path.exists(archivo_excel):
        print("❌ No se puede verificar - archivo no encontrado")
        return
    
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        hoja = libro.active
        
        print(f"📊 Filas totales después de limpiar: {hoja.max_row}")
        
        # Analizar materiales restantes
        materiales_encontrados = {}
        registros_validos = 0
        
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value
            movimiento = hoja.cell(row=row, column=5).value
            cantidad = hoja.cell(row=row, column=6).value
            usuario = hoja.cell(row=row, column=4).value
            
            if material and movimiento and cantidad:
                registros_validos += 1
                print(f"✅ {material} | {movimiento} | {cantidad} | Usuario: {usuario}")
                
                # Calcular stock
                if material not in materiales_encontrados:
                    materiales_encontrados[material] = 0
                
                try:
                    cantidad_num = float(str(cantidad).replace(",", "."))
                    if "Entrada" in str(movimiento):
                        materiales_encontrados[material] += cantidad_num
                    elif "Salida" in str(movimiento):
                        materiales_encontrados[material] -= cantidad_num
                except:
                    pass
        
        print(f"\n📊 STOCK REAL DESPUÉS DE LA LIMPIEZA:")
        for material, stock in materiales_encontrados.items():
            print(f"   📦 {material}: {stock:.1f}")
        
        print(f"\n✅ Registros válidos encontrados: {registros_validos}")
        
        # Verificar que no hay materiales falsos
        materiales_falsos_encontrados = [m for m in materiales_encontrados.keys() 
                                       if "Material_" in str(m)]
        
        if materiales_falsos_encontrados:
            print(f"⚠️ AÚN HAY MATERIALES FALSOS: {materiales_falsos_encontrados}")
            return False
        else:
            print(f"✅ NO SE ENCONTRARON MATERIALES FALSOS")
            return True
            
    except Exception as e:
        print(f"❌ Error verificando: {e}")
        return False

def main():
    """Función principal"""
    print("🧹 === LIMPIEZA DE DATOS FALSOS ===")
    print(f"⏰ Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    print("\n🎯 OBJETIVO:")
    print("• Eliminar datos de ejemplo (Material_0, Material_1, etc.)")
    print("• Eliminar registros del usuario 'Sistema'")
    print("• Conservar solo TUS datos reales")
    
    # Ejecutar limpieza
    if limpiar_datos_falsos():
        print("\n🎉 LIMPIEZA COMPLETADA")
        
        # Verificar resultado
        if verificar_resultado():
            print("\n✅ === ÉXITO ===")
            print("🎯 Tu archivo Excel ahora contiene solo datos reales")
            print("📊 La gráfica debería mostrar valores correctos")
            print("💡 Prueba generar la gráfica nuevamente en el bot")
        else:
            print("\n⚠️ === VERIFICACIÓN FALLÓ ===")
            print("Puede que queden algunos datos falsos")
    else:
        print("\n❌ === ERROR EN LA LIMPIEZA ===")
        print("No se pudo completar la limpieza")
    
    print(f"\n⏰ Finalizado: {datetime.now().strftime('%H:%M:%S')}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
