#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📊 GRAPHICS GENERATOR - VERSIÓN FINAL CORREGIDA
===============================================

Módulo para generar gráficas del Sistema Industrial Unificado
TODAS las correcciones aplicadas:
- Lee combustibles desde inventario_materiales.xlsx (NO desde equipos)
- Lee cemento desde inventario_materiales.xlsx
- Estructura de columnas corregida
- Búsqueda flexible de materiales
- Manejo robusto de errores

Autor: Sistema Industrial Automatizado
Versión: 2.0 FINAL
Fecha: 2025
"""

import os
import sys
from datetime import datetime, timedelta
import openpyxl

# Importar configuración
try:
    from .config import *
except ImportError:
    try:
        from modules.config import *
    except ImportError:
        # Configuración de respaldo
        ARCHIVO_EXCEL_MATERIALES = "datos/inventario_materiales.xlsx"
        ARCHIVO_EXCEL_EQUIPOS = "datos/inventario_equipos.xlsx"
        ARCHIVO_EXCEL_PRODUCCION = "datos/registro_produccion.xlsx"

# Verificar matplotlib
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.patches import Patch
    GRAFICOS_DISPONIBLES = True
    print("✅ Matplotlib cargado para gráficos")
except ImportError:
    GRAFICOS_DISPONIBLES = False
    print("⚠️ Matplotlib no disponible - gráficas deshabilitadas")

class GraphicsGenerator:
    """Generador de gráficas para el sistema industrial"""
    
    @staticmethod
    def verificar_matplotlib():
        """Verifica si matplotlib está disponible"""
        return GRAFICOS_DISPONIBLES
    
    @staticmethod
    def _buscar_archivo_materiales():
        """Busca el archivo de materiales en diferentes ubicaciones"""
        ubicaciones = [
            ARCHIVO_EXCEL_MATERIALES,
            "datos/inventario_materiales.xlsx",
            "inventario_materiales.xlsx",
            "./datos/inventario_materiales.xlsx"
        ]
        
        for ubicacion in ubicaciones:
            if os.path.exists(ubicacion):
                print(f"📁 Usando archivo: {ubicacion}")
                return ubicacion
        
        print("❌ No se encontró archivo de materiales")
        return None
    
    @staticmethod
    def generar_grafica_combustibles():
        """
        Genera gráfica de stock de combustibles
        CORREGIDO: Ahora lee desde inventario_materiales.xlsx
        """
        if not GRAFICOS_DISPONIBLES:
            print("❌ Matplotlib no disponible")
            return None
        
        try:
            # CORRECCIÓN: Buscar archivo de materiales (no equipos)
            archivo = GraphicsGenerator._buscar_archivo_materiales()
            if not archivo:
                return None
            
            # Leer datos del Excel
            libro = openpyxl.load_workbook(archivo)
            hoja = libro.active
            
            print(f"📊 Analizando archivo con {hoja.max_row} filas")
            
            # CORRECCIÓN: Extraer movimientos de combustible desde MATERIALES
            movimientos_gasolina = []
            movimientos_diesel = []
            
            for row in range(5, hoja.max_row + 1):
                try:
                    fecha = hoja.cell(row=row, column=1).value      # Columna A: Fecha
                    hora = hoja.cell(row=row, column=2).value       # Columna B: Hora
                    material = hoja.cell(row=row, column=3).value   # Columna C: Material
                    usuario = hoja.cell(row=row, column=4).value    # Columna D: Usuario
                    movimiento = hoja.cell(row=row, column=5).value # Columna E: Movimiento
                    cantidad = hoja.cell(row=row, column=6).value   # Columna F: Cantidad
                    
                    if material and cantidad and movimiento:
                        material_str = str(material).lower()
                        cantidad_num = float(str(cantidad).replace(",", "."))
                        
                        # CORRECCIÓN: Buscar gasolina y diesel como MATERIALES
                        if "gasolina" in material_str:
                            if "entrada" in str(movimiento).lower():
                                movimientos_gasolina.append((fecha, cantidad_num))
                            elif "salida" in str(movimiento).lower():
                                movimientos_gasolina.append((fecha, -cantidad_num))
                        
                        elif "diesel" in material_str:
                            if "entrada" in str(movimiento).lower():
                                movimientos_diesel.append((fecha, cantidad_num))
                            elif "salida" in str(movimiento).lower():
                                movimientos_diesel.append((fecha, -cantidad_num))
                                
                except Exception as e:
                    continue
            
            print(f"🔍 Movimientos gasolina encontrados: {len(movimientos_gasolina)}")
            print(f"🔍 Movimientos diesel encontrados: {len(movimientos_diesel)}")
            
            # Si no hay datos reales, usar datos de ejemplo
            if not movimientos_gasolina and not movimientos_diesel:
                print("📊 No hay datos reales, generando gráfica con datos de ejemplo")
                fechas_ejemplo = []
                stock_gasolina = [150, 130, 110, 95, 80, 65, 50]
                stock_diesel = [200, 185, 170, 155, 140, 125, 110]
                
                for i in range(7):
                    fecha = (datetime.now() - timedelta(days=6-i)).strftime("%d/%m")
                    fechas_ejemplo.append(fecha)
            else:
                # Procesar datos reales
                fechas_ejemplo = []
                stock_gasolina = []
                stock_diesel = []
                
                # Generar fechas de la última semana
                for i in range(7):
                    fecha = (datetime.now() - timedelta(days=6-i)).strftime("%d/%m/%Y")
                    fechas_ejemplo.append(fecha[-5:])  # Solo DD/MM
                    
                    # Calcular stock acumulado para cada fecha
                    acum_gasolina = 100  # Stock inicial
                    acum_diesel = 150    # Stock inicial
                    
                    for f, cantidad in movimientos_gasolina:
                        fecha_mov = str(f) if f else ""
                        if fecha in fecha_mov:
                            acum_gasolina += cantidad
                    
                    for f, cantidad in movimientos_diesel:
                        fecha_mov = str(f) if f else ""
                        if fecha in fecha_mov:
                            acum_diesel += cantidad
                    
                    stock_gasolina.append(max(0, acum_gasolina))
                    stock_diesel.append(max(0, acum_diesel))
            
            # Crear gráfica
            plt.figure(figsize=(12, 8))
            
            x_pos = range(len(fechas_ejemplo))
            
            # Barras lado a lado
            ancho_barra = 0.35
            plt.bar([x - ancho_barra/2 for x in x_pos], stock_gasolina, ancho_barra, 
                   label='Gasolina', color='#FF6B6B', alpha=0.8, edgecolor='black')
            plt.bar([x + ancho_barra/2 for x in x_pos], stock_diesel, ancho_barra, 
                   label='Diesel', color='#4ECDC4', alpha=0.8, edgecolor='black')
            
            plt.title('⛽ STOCK DE COMBUSTIBLES\nPlanta Premoldeados Tupiza', 
                     fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Fecha', fontsize=12)
            plt.ylabel('Cantidad (Litros)', fontsize=12)
            
            plt.xticks(x_pos, fechas_ejemplo, rotation=45)
            plt.legend(fontsize=11)
            plt.grid(True, alpha=0.3, linestyle='--', axis='y')
            
            # Agregar valores en las barras
            for i, (g, d) in enumerate(zip(stock_gasolina, stock_diesel)):
                plt.text(i - ancho_barra/2, g + max(stock_gasolina + stock_diesel)*0.01, 
                        f'{g:.0f}L', ha='center', va='bottom', fontsize=9, fontweight='bold')
                plt.text(i + ancho_barra/2, d + max(stock_gasolina + stock_diesel)*0.01, 
                        f'{d:.0f}L', ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            plt.tight_layout()
            
            nombre_grafica = f"combustibles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            plt.savefig(nombre_grafica, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()
            
            print(f"✅ Gráfica de combustibles generada: {nombre_grafica}")
            return nombre_grafica
            
        except Exception as e:
            print(f"❌ Error generando gráfica combustibles: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def generar_grafica_stock_materiales():
        """
        Genera gráfica de stock actual de materiales
        """
        if not GRAFICOS_DISPONIBLES:
            print("❌ Matplotlib no disponible")
            return None
        
        try:
            # Importar ExcelManager
            try:
                from .excel_manager import ExcelManager
            except ImportError:
                from modules.excel_manager import ExcelManager
            
            stock = ExcelManager.obtener_stock_materiales()
            
            if not stock:
                print("❌ No hay datos de stock para graficar")
                return None
            
            # Preparar datos
            materiales = list(stock.keys())
            cantidades = list(stock.values())
            
            # Colores según nivel de stock
            colores = []
            for cantidad in cantidades:
                if cantidad > 50:
                    colores.append('#2ECC71')  # Verde - Óptimo
                elif cantidad > 10:
                    colores.append('#F39C12')  # Naranja - Bajo
                else:
                    colores.append('#E74C3C')  # Rojo - Crítico
            
            # Crear gráfica
            plt.figure(figsize=(14, 8))
            
            barras = plt.bar(materiales, cantidades, color=colores, alpha=0.8, edgecolor='black')
            
            plt.title('📦 STOCK ACTUAL DE MATERIALES\nPlanta Premoldeados Tupiza', 
                     fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Material', fontsize=12)
            plt.ylabel('Cantidad', fontsize=12)
            plt.xticks(rotation=45, ha='right')
            plt.grid(True, alpha=0.3, linestyle='--', axis='y')
            
            # Agregar valores en las barras
            for barra, cantidad in zip(barras, cantidades):
                altura = barra.get_height()
                plt.text(barra.get_x() + barra.get_width()/2, altura + max(cantidades)*0.01,
                        f'{cantidad:.0f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
            
            # Leyenda de colores
            leyenda = [
                Patch(color='#2ECC71', label='Óptimo (>50)'),
                Patch(color='#F39C12', label='Bajo (10-50)'),
                Patch(color='#E74C3C', label='Crítico (<10)')
            ]
            plt.legend(handles=leyenda, loc='upper right')
            
            plt.tight_layout()
            
            nombre_grafica = f"stock_materiales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            plt.savefig(nombre_grafica, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()
            
            print(f"✅ Gráfica de stock generada: {nombre_grafica}")
            return nombre_grafica
            
        except Exception as e:
            print(f"❌ Error generando gráfica stock: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def generar_grafica_cemento():
        """
        Genera gráfica específica de consumo de cemento - VERSIÓN FINAL CORREGIDA
        """
        if not GRAFICOS_DISPONIBLES:
            print("❌ Matplotlib no disponible")
            return None
        
        try:
            # Buscar archivo de materiales
            archivo = GraphicsGenerator._buscar_archivo_materiales()
            if not archivo:
                return None
            
            libro = openpyxl.load_workbook(archivo)
            hoja = libro.active
            
            print(f"📊 Analizando {hoja.max_row} filas en busca de cemento...")
            
            # Diccionario para acumular consumo por fecha
            consumo_cemento = {}
            registros_encontrados = 0
            
            for row in range(5, hoja.max_row + 1):
                try:
                    fecha = hoja.cell(row=row, column=1).value      # Columna A: Fecha
                    material = hoja.cell(row=row, column=3).value   # Columna C: Material
                    movimiento = hoja.cell(row=row, column=5).value # Columna E: Movimiento
                    cantidad = hoja.cell(row=row, column=6).value   # Columna F: Cantidad
                    
                    # CORRECCIÓN: Búsqueda flexible de cemento
                    if material and "cemento" in str(material).lower():
                        registros_encontrados += 1
                        print(f"   📦 Cemento encontrado en fila {row}: {material} | {movimiento} | {cantidad}")
                        
                        # Solo procesar salidas (consumo)
                        if movimiento and "salida" in str(movimiento).lower():
                            
                            # Procesar fecha
                            if isinstance(fecha, datetime):
                                fecha_str = fecha.strftime("%d/%m")
                            else:
                                fecha_str = str(fecha)[-5:] if fecha else "S/F"  # Últimos 5 chars (DD/MM)
                            
                            # Procesar cantidad
                            try:
                                cantidad_num = float(str(cantidad).replace(",", "."))
                                
                                if fecha_str not in consumo_cemento:
                                    consumo_cemento[fecha_str] = 0
                                
                                consumo_cemento[fecha_str] += cantidad_num
                                print(f"      ✅ Consumo registrado: {fecha_str} = {cantidad_num} bolsas")
                                
                            except (ValueError, TypeError):
                                print(f"      ⚠️ Cantidad inválida: {cantidad}")
                                continue
                        else:
                            print(f"      ⚠️ No es salida: {movimiento}")
                            
                except Exception as e:
                    continue
            
            print(f"📊 Registros de cemento encontrados: {registros_encontrados}")
            print(f"📊 Días con consumo: {len(consumo_cemento)}")
            
            # Verificar si hay datos de consumo
            if not consumo_cemento:
                print("❌ No hay datos de consumo de cemento (salidas)")
                print("💡 Para generar la gráfica necesitas registrar SALIDAS de cemento")
                return None
            
            # Si hay muy pocos datos, agregar algunos días de ejemplo
            if len(consumo_cemento) < 3:
                print("⚠️ Pocos datos reales, agregando días de ejemplo...")
                for i in range(1, 6):
                    fecha_ej = (datetime.now() - timedelta(days=i)).strftime("%d/%m")
                    if fecha_ej not in consumo_cemento:
                        consumo_cemento[fecha_ej] = 0
            
            # Preparar datos para gráfica
            fechas = sorted(consumo_cemento.keys())
            cantidades = [consumo_cemento[f] for f in fechas]
            
            # Crear gráfica
            plt.figure(figsize=(12, 8))
            
            barras = plt.bar(range(len(fechas)), cantidades, 
                           color='#8E44AD', alpha=0.8, edgecolor='black', linewidth=1)
            
            plt.title('🏗️ CONSUMO DIARIO DE CEMENTO\nPlanta Premoldeados Tupiza', 
                     fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Fecha', fontsize=12)
            plt.ylabel('Bolsas Consumidas', fontsize=12)
            
            plt.xticks(range(len(fechas)), fechas, rotation=45)
            plt.grid(True, alpha=0.3, linestyle='--', axis='y')
            
            # Agregar valores en las barras
            for i, (barra, cantidad) in enumerate(zip(barras, cantidades)):
                if cantidad > 0:  # Solo mostrar si hay consumo
                    plt.text(barra.get_x() + barra.get_width()/2, 
                            barra.get_height() + max(cantidades)*0.02,
                            f'{cantidad:.0f}', ha='center', va='bottom', 
                            fontweight='bold', fontsize=11)
            
            # Información adicional
            total_consumo = sum(cantidades)
            dias_con_consumo = sum(1 for c in cantidades if c > 0)
            promedio_diario = total_consumo / dias_con_consumo if dias_con_consumo > 0 else 0
            
            # CORRECCIÓN: F-string arreglado
            info_text = f'Total: {total_consumo:.0f} bolsas | Promedio: {promedio_diario:.1f} bolsas/día | Días activos: {dias_con_consumo}'
            plt.figtext(0.02, 0.02, info_text, fontsize=10, style='italic')
            
            plt.tight_layout()
            
            nombre_grafica = f"cemento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            plt.savefig(nombre_grafica, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()
            
            print(f"✅ Gráfica de cemento generada: {nombre_grafica}")
            return nombre_grafica
            
        except Exception as e:
            print(f"❌ Error generando gráfica cemento: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def generar_todas_las_graficas():
        """Genera todas las gráficas disponibles"""
        if not GRAFICOS_DISPONIBLES:
            print("❌ Matplotlib no disponible")
            return []
        
        graficas_generadas = []
        
        print("📊 === GENERANDO TODAS LAS GRÁFICAS ===")
        
        # Gráfica de combustibles
        print("\n⛽ Generando gráfica de combustibles...")
        try:
            grafica1 = GraphicsGenerator.generar_grafica_combustibles()
            if grafica1:
                graficas_generadas.append(grafica1)
        except Exception as e:
            print(f"❌ Error en gráfica combustibles: {e}")
        
        # Gráfica de stock de materiales
        print("\n📦 Generando gráfica de stock de materiales...")
        try:
            grafica2 = GraphicsGenerator.generar_grafica_stock_materiales()
            if grafica2:
                graficas_generadas.append(grafica2)
        except Exception as e:
            print(f"❌ Error en gráfica stock: {e}")
        
        # Gráfica de cemento
        print("\n🏗️ Generando gráfica de cemento...")
        try:
            grafica3 = GraphicsGenerator.generar_grafica_cemento()
            if grafica3:
                graficas_generadas.append(grafica3)
        except Exception as e:
            print(f"❌ Error en gráfica cemento: {e}")
        
        print(f"\n📈 === RESUMEN ===")
        print(f"✅ Gráficas generadas: {len(graficas_generadas)}/3")
        for i, grafica in enumerate(graficas_generadas, 1):
            print(f"   {i}. {grafica}")
        
        return graficas_generadas
    
    @staticmethod
    def limpiar_graficas_antiguas(dias=7):
        """Limpia gráficas más antiguas que X días"""
        try:
            import glob
            import time
            
            patrones = ["combustibles_*.png", "cemento_*.png", "stock_*.png"]
            eliminados = 0
            
            for patron in patrones:
                archivos = glob.glob(patron)
                for archivo in archivos:
                    if os.path.getctime(archivo) < time.time() - (dias * 24 * 60 * 60):
                        os.remove(archivo)
                        eliminados += 1
            
            if eliminados > 0:
                print(f"🧹 Limpieza completada: {eliminados} archivos eliminados")
            return eliminados
            
        except Exception as e:
            print(f"❌ Error en limpieza: {e}")
            return 0

# ============================================================================
# FUNCIONES DE UTILIDAD Y PRUEBAS
# ============================================================================

def probar_graphics_generator():
    """Prueba todas las funciones del generador de gráficas"""
    print("🧪 === PRUEBA COMPLETA DE GRAPHICS GENERATOR ===")
    print(f"⏰ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    if not GraphicsGenerator.verificar_matplotlib():
        print("❌ Matplotlib no disponible - no se pueden generar gráficas")
        return False
    
    # Verificar archivo de datos
    archivo = GraphicsGenerator._buscar_archivo_materiales()
    if not archivo:
        print("❌ No se encontró archivo de materiales")
        return False
    
    # Probar cada función
    funciones = [
        ("Combustibles", GraphicsGenerator.generar_grafica_combustibles),
        ("Stock Materiales", GraphicsGenerator.generar_grafica_stock_materiales),
        ("Cemento", GraphicsGenerator.generar_grafica_cemento)
    ]
    
    resultados = []
    
    for nombre, funcion in funciones:
        try:
            print(f"\n📊 Probando gráfica de {nombre}...")
            resultado = funcion()
            
            if resultado:
                print(f"   ✅ Generada: {resultado}")
                resultados.append(resultado)
                
                # Verificar archivo
                if os.path.exists(resultado):
                    tamaño = os.path.getsize(resultado) / 1024
                    print(f"   📏 Tamaño: {tamaño:.1f} KB")
            else:
                print(f"   ⚠️ No se generó (revisar datos)")
                
        except Exception as e:
            print(f"   ❌ Error: {e}")
    
    print(f"\n📈 === RESUMEN FINAL ===")
    print(f"✅ Gráficas exitosas: {len(resultados)}/{len(funciones)}")
    
    # Limpiar archivos de prueba después de 5 segundos
    import time
    time.sleep(2)
    for archivo in resultados:
        try:
            if os.path.exists(archivo):
                print(f"🧹 Limpiando: {archivo}")
                os.remove(archivo)
        except:
            pass
    
    return len(resultados) > 0

if __name__ == "__main__":
    probar_graphics_generator()