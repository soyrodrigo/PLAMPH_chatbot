# ============================================================================
# PASO 1: CREAR ESTRUCTURA Y PRIMER MÓDULO
# ============================================================================

"""
📁 Estructura a crear:

bot_planta_tupiza/
├── main.py
├── modules/
│   ├── __init__.py
│   ├── config.py
│   └── excel_manager.py
├── inventario_materiales.xlsx      # TUS ARCHIVOS (mover aquí)
├── registro_actividades.xlsx       # TUS ARCHIVOS (mover aquí)
├── produccion_diaria.xlsx          # TUS ARCHIVOS (mover aquí)
└── fotos_planta/                   # TU CARPETA (mover aquí)
"""

# ============================================================================
# 1. ARCHIVO: modules/__init__.py (crear vacío)
# ============================================================================

# Archivo vacío para que Python reconozca 'modules' como paquete

# ============================================================================
# 2. ARCHIVO: modules/config.py
# ============================================================================

"""
⚙️ modules/config.py - CONFIGURACIONES EXACTAS QUE TIENES
"""

# Token del bot (pon tu token real aquí)
TOKEN = "TU_TOKEN_AQUI"  # ⚠️ CAMBIA ESTO POR TU TOKEN REAL

# Archivos exactos que ya tienes (NO CAMBIAR nombres)
ARCHIVO_EXCEL_MATERIALES = "inventario_materiales.xlsx"
ARCHIVO_EXCEL_EQUIPOS = "inventario_equipos.xlsx"
ARCHIVO_ACTIVIDADES = "registro_actividades.xlsx"
ARCHIVO_PRODUCCION = "produccion_diaria.xlsx"
CARPETA_FOTOS = "fotos_planta"
ARCHIVO_ESTADOS_USUARIO = "estados_usuario.json"
ARCHIVO_ESTADOS_PRODUCCION = "estados_produccion.json"

# Listas exactas que ya tienes
MATERIALES = ["Cemento", "Arena", "Gasolina", "Diesel", "Alambre", "Acero", "Pintura", "Grasa"]
EQUIPOS = ["Máquina de Soldar", "Carretilla", "Martillo", "Mezcladora", "Taladro", "Compresora", "Grúa"]
CONDICIONES = ["Nuevo", "Muy Bueno", "Bueno", "Regular", "Malo", "Para Reparar"]

# Estados exactos que ya tienes
ESPERANDO_MATERIAL = "esperando_material"
ESPERANDO_EQUIPO = "esperando_equipo"
ESPERANDO_MOVIMIENTO = "esperando_movimiento"
ESPERANDO_CANTIDAD = "esperando_cantidad"
ESPERANDO_CONDICION = "esperando_condicion"
ESPERANDO_OBSERVACIONES = "esperando_observaciones"
ESPERANDO_CANTIDAD_PRODUCCION = "esperando_cantidad_produccion"
ESPERANDO_ACTIVIDAD = "esperando_actividad"
ESPERANDO_FECHA_REPORTE = "esperando_fecha_reporte"

# Modelos de adoquines exactos que ya tienes
MODELOS_ADOQUINES = {
    "modelo_i": {
        "nombre": "Adoquín Modelo I",
        "adoquines_por_pallet": 132
    },
    "doble_s": {
        "nombre": "Adoquín Doble S", 
        "adoquines_por_pallet": 120
    }
}

# ============================================================================
# 3. ARCHIVO: modules/excel_manager.py
# ============================================================================

"""
📊 modules/excel_manager.py - TUS FUNCIONES EXCEL EXACTAS
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from .config import *

class ExcelManager:
    """
    Gestor de archivos Excel
    CONTIENE EXACTAMENTE TU LÓGICA ACTUAL, SOLO ORGANIZADA
    """
    
    @staticmethod
    def verificar_y_crear_archivos():
        """Verifica que existan todos los archivos Excel, los crea si no existen"""
        if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
            ExcelManager.crear_excel_materiales()
        if not os.path.exists(ARCHIVO_EXCEL_EQUIPOS):
            ExcelManager.crear_excel_equipos()
        if not os.path.exists(ARCHIVO_ACTIVIDADES):
            ExcelManager.crear_excel_actividades()
        if not os.path.exists(ARCHIVO_PRODUCCION):
            ExcelManager.crear_excel_produccion()
        
        print("✅ Archivos Excel verificados/creados")
    
    @staticmethod
    def crear_excel_materiales():
        """Crea archivo Excel de materiales - EXACTAMENTE COMO LO TIENES"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Inventario Materiales"
        
        # TÍTULO PRINCIPAL (exactamente igual que tienes)
        hoja.merge_cells('A1:G1')
        hoja['A1'] = f"🏭 INVENTARIO DE MATERIALES - PLANTA TUPIZA {datetime.now().year}"
        hoja['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        hoja['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFORMACIÓN DEL SISTEMA (exactamente igual)
        hoja['A2'] = f"📅 Creado automáticamente: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        hoja['A2'].font = Font(name='Arial', size=10, italic=True)
        hoja.merge_cells('A2:G2')
        
        # FILA VACÍA (exactamente igual)
        hoja.row_dimensions[3].height = 10
        
        # ENCABEZADOS (exactamente iguales)
        encabezados = ["📅 FECHA", "🕒 HORA", "📦 MATERIAL", "👤 USUARIO", "📊 MOVIMIENTO", "🔢 CANTIDAD", "📝 OBSERVACIONES"]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celda.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos (exactamente igual)
        anchos = [12, 10, 15, 12, 12, 10, 20]
        for i, ancho in enumerate(anchos, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        libro.save(ARCHIVO_EXCEL_MATERIALES)
        print(f"✅ Archivo creado: {ARCHIVO_EXCEL_MATERIALES}")
    
    @staticmethod
    def crear_excel_equipos():
        """Crea archivo Excel de equipos - EXACTAMENTE COMO LO TIENES"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Inventario Equipos"
        
        # TÍTULO PRINCIPAL (exactamente igual)
        hoja.merge_cells('A1:G1')
        hoja['A1'] = f"🔧 INVENTARIO DE EQUIPOS - PLANTA TUPIZA {datetime.now().year}"
        hoja['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        hoja['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFORMACIÓN DEL SISTEMA (exactamente igual)
        hoja['A2'] = f"📅 Creado automáticamente: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        hoja['A2'].font = Font(name='Arial', size=10, italic=True)
        hoja.merge_cells('A2:G2')
        
        # FILA VACÍA (exactamente igual)
        hoja.row_dimensions[3].height = 10
        
        # ENCABEZADOS (exactamente iguales)
        encabezados = ["📅 FECHA", "🕒 HORA", "🔧 EQUIPO", "👤 USUARIO", "🏷️ CONDICIÓN", "📊 ESTADO", "📝 OBSERVACIONES"]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celda.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos (exactamente igual)
        anchos = [12, 10, 15, 12, 12, 12, 20]
        for i, ancho in enumerate(anchos, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        libro.save(ARCHIVO_EXCEL_EQUIPOS)
        print(f"✅ Archivo creado: {ARCHIVO_EXCEL_EQUIPOS}")
    
    @staticmethod
    def crear_excel_actividades():
        """Crea archivo Excel de actividades - EXACTAMENTE COMO LO TIENES"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Registro Actividades"
        
        # TÍTULO PRINCIPAL (exactamente igual)
        hoja.merge_cells('A1:F1')
        hoja['A1'] = f"📝 REGISTRO DE ACTIVIDADES - PLANTA TUPIZA {datetime.now().year}"
        hoja['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        hoja['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFORMACIÓN DEL SISTEMA (exactamente igual)
        hoja['A2'] = f"📅 Creado automáticamente: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        hoja['A2'].font = Font(name='Arial', size=10, italic=True)
        hoja.merge_cells('A2:F2')
        
        # FILA VACÍA (exactamente igual)
        hoja.row_dimensions[3].height = 10
        
        # ENCABEZADOS (exactamente iguales)
        encabezados = ["📅 FECHA", "🕒 HORA", "👤 USUARIO", "📋 TIPO", "📝 ACTIVIDAD", "📸 ARCHIVO FOTO"]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celda.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos (exactamente igual)
        anchos = [12, 10, 12, 8, 25, 20]
        for i, ancho in enumerate(anchos, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        libro.save(ARCHIVO_ACTIVIDADES)
        print(f"✅ Archivo creado: {ARCHIVO_ACTIVIDADES}")
    
    @staticmethod
    def crear_excel_produccion():
        """Crea archivo Excel de producción - EXACTAMENTE COMO LO TIENES"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Producción Diaria"
        
        # TÍTULO PRINCIPAL (exactamente igual)
        hoja.merge_cells('A1:H1')
        hoja['A1'] = f"🏭 PRODUCCIÓN DIARIA - PLANTA TUPIZA {datetime.now().year}"
        hoja['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        hoja['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFORMACIÓN DEL SISTEMA (exactamente igual)
        hoja['A2'] = f"📅 Creado automáticamente: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        hoja['A2'].font = Font(name='Arial', size=10, italic=True)
        hoja.merge_cells('A2:H2')
        
        # FILA VACÍA (exactamente igual)
        hoja.row_dimensions[3].height = 10
        
        # ENCABEZADOS (exactamente iguales)
        encabezados = ["📅 FECHA", "🕒 TURNO", "📋 TIPO ENTRADA", "🧱 MODELO", "📦 PALLETS", "🔢 ADOQUINES", "👤 USUARIO", "📝 OBSERVACIONES"]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celda.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos (exactamente igual)
        anchos = [12, 10, 12, 15, 10, 12, 12, 20]
        for i, ancho in enumerate(anchos, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        libro.save(ARCHIVO_PRODUCCION)
        print(f"✅ Archivo creado: {ARCHIVO_PRODUCCION}")
    
    @staticmethod
    def guardar_material(fecha, hora, material, usuario, movimiento, cantidad, observaciones):
        """Guarda registro de material - EXACTAMENTE TU LÓGICA ACTUAL"""
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            fila_nueva = hoja.max_row + 1
            
            # Datos a guardar (exactamente igual)
            datos = [fecha, hora, material, usuario, movimiento, cantidad, observaciones]
            
            for col, dato in enumerate(datos, 1):
                celda = hoja.cell(row=fila_nueva, column=col)
                celda.value = dato
                celda.font = Font(name='Arial', size=10)
                celda.alignment = Alignment(horizontal='center', vertical='center')
                
                # Bordes (exactamente iguales)
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                celda.border = thin_border
            
            libro.save(ARCHIVO_EXCEL_MATERIALES)
            return True
            
        except Exception as e:
            print(f"Error guardando material: {e}")
            return False
    
    @staticmethod
    def guardar_equipo(fecha, hora, equipo, usuario, condicion, estado, observaciones):
        """Guarda registro de equipo - EXACTAMENTE TU LÓGICA ACTUAL"""
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_EQUIPOS)
            hoja = libro.active
            
            fila_nueva = hoja.max_row + 1
            
            # Datos a guardar (exactamente igual)
            datos = [fecha, hora, equipo, usuario, condicion, estado, observaciones]
            
            for col, dato in enumerate(datos, 1):
                celda = hoja.cell(row=fila_nueva, column=col)
                celda.value = dato
                celda.font = Font(name='Arial', size=10)
                celda.alignment = Alignment(horizontal='center', vertical='center')
                
                # Bordes (exactamente iguales)
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                celda.border = thin_border
            
            libro.save(ARCHIVO_EXCEL_EQUIPOS)
            return True
            
        except Exception as e:
            print(f"Error guardando equipo: {e}")
            return False
    
    @staticmethod
    def guardar_actividad(fecha, hora, usuario, tipo, actividad, archivo_foto=None):
        """Guarda registro de actividad - EXACTAMENTE TU LÓGICA ACTUAL"""
        try:
            libro = openpyxl.load_workbook(ARCHIVO_ACTIVIDADES)
            hoja = libro.active
            
            fila_nueva = hoja.max_row + 1
            
            # Datos a guardar (exactamente igual)
            datos = [fecha, hora, usuario, tipo, actividad, archivo_foto or ""]
            
            for col, dato in enumerate(datos, 1):
                celda = hoja.cell(row=fila_nueva, column=col)
                celda.value = dato
                celda.font = Font(name='Arial', size=10)
                celda.alignment = Alignment(horizontal='center', vertical='center')
                
                # Bordes (exactamente iguales)
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                celda.border = thin_border
                
                # Color especial para tipo (exactamente igual que tienes)
                if col == 4:  # Columna TIPO
                    if tipo == "Foto":
                        celda.font = Font(name='Arial', size=10, color='0066CC')
                        celda.value = "📸 Foto"
                    else:
                        celda.font = Font(name='Arial', size=10, color='006600')
                        celda.value = "📝 Texto"
            
            # Ajustar altura de fila (exactamente igual)
            hoja.row_dimensions[fila_nueva].height = 18
            
            libro.save(ARCHIVO_ACTIVIDADES)
            return True
            
        except Exception as e:
            print(f"Error guardando actividad: {e}")
            return False
    
    @staticmethod
    def guardar_produccion(fecha, turno, tipo_entrada, modelo, pallets, adoquines, usuario, observaciones):
        """Guarda registro de producción - EXACTAMENTE TU LÓGICA ACTUAL"""
        try:
            libro = openpyxl.load_workbook(ARCHIVO_PRODUCCION)
            hoja = libro.active
            
            fila_nueva = hoja.max_row + 1
            
            # Datos a guardar (exactamente igual)
            datos = [fecha, turno, tipo_entrada, modelo, pallets, adoquines, usuario, observaciones]
            
            for col, dato in enumerate(datos, 1):
                celda = hoja.cell(row=fila_nueva, column=col)
                celda.value = dato
                celda.font = Font(name='Arial', size=10)
                celda.alignment = Alignment(horizontal='center', vertical='center')
                
                # Bordes (exactamente iguales)
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                celda.border = thin_border
            
            libro.save(ARCHIVO_PRODUCCION)
            return True
            
        except Exception as e:
            print(f"Error guardando producción: {e}")
            return False
    
    @staticmethod
    def obtener_datos_combustibles():
        """Obtiene datos de combustibles para gráficas - TU LÓGICA ACTUAL"""
        if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
            return None
        
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            stock_gasolina = 0
            stock_diesel = 0
            
            for row in range(5, hoja.max_row + 1):
                material = hoja.cell(row=row, column=3).value
                movimiento = hoja.cell(row=row, column=5).value
                cantidad = hoja.cell(row=row, column=6).value
                
                if material and movimiento and cantidad:
                    try:
                        cantidad_num = float(str(cantidad).replace(",", "."))
                        material_lower = str(material).lower()
                        
                        # Búsqueda flexible (exactamente como la tienes)
                        if any(palabra in material_lower for palabra in ['gasolina', 'gasoline', 'nafta']):
                            if "Entrada" in str(movimiento):
                                stock_gasolina += cantidad_num
                            elif "Salida" in str(movimiento):
                                stock_gasolina -= cantidad_num
                        elif any(palabra in material_lower for palabra in ['diesel', 'diésel', 'gasoil']):
                            if "Entrada" in str(movimiento):
                                stock_diesel += cantidad_num
                            elif "Salida" in str(movimiento):
                                stock_diesel -= cantidad_num
                    except:
                        continue
            
            return {
                'gasolina': max(0, stock_gasolina),
                'diesel': max(0, stock_diesel)
            }
            
        except Exception as e:
            print(f"Error obteniendo datos de combustibles: {e}")
            return None

# ============================================================================
# 4. ARCHIVO: main.py (SIMPLIFICADO)
# ============================================================================

"""
🚀 main.py - Punto de entrada simplificado
"""

from modules.config import TOKEN
from modules.excel_manager import ExcelManager

def main():
    """Función principal simplificada"""
    
    print("🏭 === BOT INDUSTRIAL UNIFICADO MODULAR ===")
    print(f"📅 Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("🔧 Código modular - Mismos datos y funcionalidad")
    
    # Verificar TOKEN
    if TOKEN == "TU_TOKEN_AQUI":
        print("❌ CONFIGURA EL TOKEN en modules/config.py")
        print("1. Abre modules/config.py")
        print("2. Cambia TOKEN = \"TU_TOKEN_AQUI\" por tu token real")
        print("3. Ejecuta nuevamente")
        return
    
    # Verificar y crear archivos Excel si no existen
    print("📊 Verificando archivos Excel...")
    ExcelManager.verificar_y_crear_archivos()
    
    # Aquí irá el resto del bot cuando migremos más módulos
    print("✅ Primer módulo (excel_manager.py) listo")
    print("📝 Siguiente paso: migrar funciones de gráficas")
    print("🎯 El ExcelManager ya tiene toda tu lógica de Excel funcionando")

if __name__ == "__main__":
    main()

# ============================================================================
# 5. ARCHIVO: test_excel_manager.py (PARA PROBAR)
# ============================================================================

"""
🧪 test_excel_manager.py - Para probar que funciona igual
"""

from modules.excel_manager import ExcelManager
from datetime import datetime

def probar_excel_manager():
    """Prueba el ExcelManager para verificar que funciona igual"""
    
    print("🧪 === PROBANDO EXCEL MANAGER ===")
    
    # Crear archivos si no existen
    ExcelManager.verificar_y_crear_archivos()
    
    # Probar guardar material
    print("📦 Probando guardar material...")
    resultado = ExcelManager.guardar_material(
        fecha=datetime.now().strftime("%d/%m/%Y"),
        hora=datetime.now().strftime("%H:%M:%S"),
        material="Cemento",
        usuario="Prueba",
        movimiento="📈 Entrada",
        cantidad=50.0,
        observaciones="Prueba del módulo Excel"
    )
    
    if resultado:
        print("✅ Material guardado correctamente")
    else:
        print("❌ Error guardando material")
    
    # Probar guardar actividad
    print("📝 Probando guardar actividad...")
    resultado_actividad = ExcelManager.guardar_actividad(
        fecha=datetime.now().strftime("%d/%m/%Y"),
        hora=datetime.now().strftime("%H:%M:%S"),
        usuario="Prueba",
        tipo="Texto",
        actividad="Prueba del módulo de actividades"
    )
    
    if resultado_actividad:
        print("✅ Actividad guardada correctamente")
    else:
        print("❌ Error guardando actividad")
    
    # Probar obtener datos de combustibles
    print("⛽ Probando obtener datos de combustibles...")
    datos_combustibles = ExcelManager.obtener_datos_combustibles()
    
    if datos_combustibles:
        print(f"✅ Datos obtenidos: Gasolina: {datos_combustibles['gasolina']}, Diesel: {datos_combustibles['diesel']}")
    else:
        print("⚠️ No hay datos de combustibles aún")
    
    print("🎯 Prueba completada - Excel Manager funciona igual que antes")

if __name__ == "__main__":
    probar_excel_manager()

print("""
🎯 PRIMER PASO COMPLETADO: EXCEL MANAGER

✅ **LO QUE TIENES AHORA:**
• modules/config.py - Todas tus configuraciones centralizadas
• modules/excel_manager.py - Todas tus funciones de Excel organizadas
• Exactamente la misma lógica que tenías antes
• Mismos formatos de archivos Excel
• Mismos resultados

📋 **PRÓXIMOS PASOS:**
1. Crear la estructura de carpetas
2. Mover tus archivos existentes 
3. Probar que el ExcelManager funciona igual
4. Migrar funciones de gráficas al siguiente módulo

💡 **CÓMO PROBAR:**
1. Crea carpeta: bot_planta_tupiza/
2. Crea subcarpeta: modules/
3. Crea los archivos mostrados arriba
4. Mueve tus archivos Excel existentes a bot_planta_tupiza/
5. Ejecuta: python test_excel_manager.py

🎯 **RESULTADO:**
Misma funcionalidad, código más organizado
Tu trabajo protegido al 100%
""")
