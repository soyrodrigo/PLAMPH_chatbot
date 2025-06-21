"""
📊 modules/excel_manager.py - GESTIÓN DE ARCHIVOS EXCEL
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from .config import *

class ExcelManager:
    """
    Gestor de archivos Excel
    CONTIENE EXACTAMENTE TU LÓGICA ACTUAL, SOLO ORGANIZADA
    """
    
    @staticmethod
    def verificar_y_crear_archivos():
        """Verifica y crea archivos Excel si no existen - EXACTAMENTE TU LÓGICA"""
        archivos = [
            (ARCHIVO_EXCEL_MATERIALES, ExcelManager.crear_estructura_materiales),
            (ARCHIVO_EXCEL_EQUIPOS, ExcelManager.crear_estructura_equipos),
            (ARCHIVO_EXCEL_PRODUCCION, ExcelManager.crear_estructura_produccion)
        ]
        
        for archivo, funcion_crear in archivos:
            if not os.path.exists(archivo):
                print(f"📄 Creando archivo: {archivo}")
                funcion_crear(archivo)
            else:
                print(f"✅ Archivo existe: {archivo}")
    
    @staticmethod
    def crear_estructura_materiales(archivo):
        """Crea estructura del archivo de materiales - EXACTAMENTE TU LÓGICA"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Inventario Materiales"
        
        # Encabezados (exactamente como los tienes)
        encabezados = ["Fecha", "Hora", "Material", "Proveedor/Destino", "Tipo Movimiento", "Cantidad", "Observaciones"]
        
        # Estilos
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = font_encabezado
            celda.fill = fill_encabezado
            celda.alignment = Alignment(horizontal="center")
        
        # Título
        hoja.merge_cells("A1:G3")
        titulo = hoja["A1"]
        titulo.value = "INVENTARIO DE MATERIALES - PLANTA PREMOLDEADOS TUPIZA"
        titulo.font = Font(bold=True, size=14, color="366092")
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        
        libro.save(archivo)
        print(f"✅ Estructura de materiales creada: {archivo}")
    
    @staticmethod
    def crear_estructura_equipos(archivo):
        """Crea estructura del archivo de equipos"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Inventario Equipos"
        
        encabezados = ["Fecha", "Código", "Nombre", "Tipo", "Condición", "Ubicación", "Observaciones"]
        
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = font_encabezado
            celda.fill = fill_encabezado
            celda.alignment = Alignment(horizontal="center")
        
        hoja.merge_cells("A1:G3")
        titulo = hoja["A1"]
        titulo.value = "INVENTARIO DE EQUIPOS - PLANTA PREMOLDEADOS TUPIZA"
        titulo.font = Font(bold=True, size=14, color="92D050")
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        
        libro.save(archivo)
        print(f"✅ Estructura de equipos creada: {archivo}")
    
    @staticmethod
    def crear_estructura_produccion(archivo):
        """Crea estructura del archivo de producción"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Registro Producción"
        
        encabezados = ["Fecha", "Hora", "Producto", "Cantidad", "Operador", "Observaciones"]
        
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = font_encabezado
            celda.fill = fill_encabezado
            celda.alignment = Alignment(horizontal="center")
        
        hoja.merge_cells("A1:F3")
        titulo = hoja["A1"]
        titulo.value = "REGISTRO DE PRODUCCIÓN - PLANTA PREMOLDEADOS TUPIZA"
        titulo.font = Font(bold=True, size=14, color="FFC000")
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        
        libro.save(archivo)
        print(f"✅ Estructura de producción creada: {archivo}")
    
    @staticmethod
    def guardar_material(fecha, hora, material, proveedor, tipo_movimiento, cantidad, observaciones):
        """Guarda un material en Excel - EXACTAMENTE TU LÓGICA ACTUAL"""
        try:
            if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
                ExcelManager.crear_estructura_materiales(ARCHIVO_EXCEL_MATERIALES)
            
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            # Encontrar la próxima fila disponible
            fila = 5
            while hoja.cell(row=fila, column=1).value is not None:
                fila += 1
            
            # Guardar datos
            datos = [fecha, hora, material, proveedor, tipo_movimiento, cantidad, observaciones]
            
            for col, dato in enumerate(datos, 1):
                hoja.cell(row=fila, column=col, value=dato)
            
            libro.save(ARCHIVO_EXCEL_MATERIALES)
            return True
            
        except Exception as e:
            print(f"Error guardando material: {e}")
            return False
    
    @staticmethod
    def obtener_stock_materiales():
        """Obtiene el stock actual de materiales - EXACTAMENTE TU LÓGICA"""
        if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
            return {}
        
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            stock = {}
            
            # Leer todos los movimientos
            for fila in range(5, hoja.max_row + 1):
                material = hoja.cell(row=fila, column=3).value
                tipo_movimiento = hoja.cell(row=fila, column=5).value
                cantidad = hoja.cell(row=fila, column=6).value
                
                if material and tipo_movimiento and cantidad:
                    try:
                        cantidad_num = float(str(cantidad).replace(",", "."))
                        
                        if material not in stock:
                            stock[material] = 0
                        
                        if "Entrada" in str(tipo_movimiento):
                            stock[material] += cantidad_num
                        elif "Salida" in str(tipo_movimiento):
                            stock[material] -= cantidad_num
                            
                    except (ValueError, TypeError):
                        continue
            
            return stock
            
        except Exception as e:
            print(f"Error obteniendo stock: {e}")
            return {}
    
    @staticmethod
    def obtener_datos_combustibles():
        """Obtiene datos específicos de combustibles - EXACTAMENTE TU LÓGICA"""
        stock = ExcelManager.obtener_stock_materiales()
        
        combustibles = {}
        
        # Buscar gasolina y diesel en el stock
        for material, cantidad in stock.items():
            material_lower = material.lower()
            if "gasolina" in material_lower:
                combustibles["gasolina"] = cantidad
            elif "diesel" in material_lower:
                combustibles["diesel"] = cantidad
        
        # Si no hay datos, devolver valores por defecto
        if not combustibles:
            combustibles = {
                "gasolina": 40.0,
                "diesel": 70.0
            }
        
        return combustibles
    
    @staticmethod
    def obtener_ultimos_movimientos(cantidad=10):
        """Obtiene los últimos movimientos registrados"""
        if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
            return []
        
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            movimientos = []
            
            # Leer desde la última fila hacia arriba
            for fila in range(max(5, hoja.max_row - cantidad + 1), hoja.max_row + 1):
                fecha = hoja.cell(row=fila, column=1).value
                hora = hoja.cell(row=fila, column=2).value
                material = hoja.cell(row=fila, column=3).value
                proveedor = hoja.cell(row=fila, column=4).value
                tipo = hoja.cell(row=fila, column=5).value
                cantidad = hoja.cell(row=fila, column=6).value
                observaciones = hoja.cell(row=fila, column=7).value
                
                if material and tipo and cantidad:
                    movimientos.append({
                        "fecha": str(fecha) if fecha else "",
                        "hora": str(hora) if hora else "",
                        "material": str(material),
                        "proveedor": str(proveedor) if proveedor else "",
                        "tipo": str(tipo),
                        "cantidad": float(str(cantidad).replace(",", ".")) if cantidad else 0,
                        "observaciones": str(observaciones) if observaciones else ""
                    })
            
            return list(reversed(movimientos))  # Más recientes primero
            
        except Exception as e:
            print(f"Error obteniendo movimientos: {e}")
            return []
    
    @staticmethod
    def contar_registros_materiales():
        """Cuenta el total de registros en materiales"""
        if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
            return 0
        
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
            hoja = libro.active
            
            # Contar filas con datos (empezando desde fila 5)
            contador = 0
            for fila in range(5, hoja.max_row + 1):
                if hoja.cell(row=fila, column=3).value:  # Si hay material
                    contador += 1
            
            return contador
            
        except Exception as e:
            print(f"Error contando registros: {e}")
            return 0