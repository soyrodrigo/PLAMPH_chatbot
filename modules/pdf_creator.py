"""
📄 modules/pdf_creator.py - GENERACIÓN DE REPORTES PDF PROFESIONALES CON ENCABEZADO Y PIE
Versión mejorada con encabezado institucional con imagen y pie de página personalizado
"""

import os
from datetime import datetime
from .config import *
from .excel_manager import ExcelManager

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import PageTemplate, Frame, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus.flowables import Flowable
    PDF_DISPONIBLE = True
    print("✅ ReportLab cargado correctamente")
except ImportError as e:
    PDF_DISPONIBLE = False
    print(f"⚠️ ReportLab no disponible: {e}")
    print("💡 Instala con: pip install reportlab")

class EncabezadoPersonalizado:
    """Clase para agregar encabezado con imagen y pie de página en todas las páginas del PDF
    SOLUCIÓN A PROBLEMAS COMUNES:
    - ✅ Encabezado y pie en TODAS las páginas (no solo la primera)
    - ✅ MISMO tamaño en primera página y páginas siguientes 
    - ✅ Posiciones exactas y consistentes
    """
    
    def __init__(self, ruta_imagen="encabezado_tupiza.png", ruta_pie="pie_tupiza.png"):
        self.ruta_imagen = ruta_imagen
        self.ruta_pie = ruta_pie
        # CONSTANTES para asegurar MISMO tamaño siempre
        self.ENCABEZADO_X = 72
        self.ENCABEZADO_WIDTH = 7*inch
        self.ENCABEZADO_HEIGHT = 1.2*inch
        self.PIE_X = 72
        self.PIE_Y = 25
        self.PIE_WIDTH = 7*inch
        self.PIE_HEIGHT = 0.8*inch
    
    def primera_pagina(self, canvas, doc):
        """Encabezado y pie para la primera página - MISMAS dimensiones que el resto"""
        canvas.saveState()
        try:
            # ==========================================
            # ENCABEZADO (parte superior) - COORDENADAS FIJAS
            # ==========================================
            if os.path.exists(self.ruta_imagen):
                canvas.drawImage(self.ruta_imagen, 
                               self.ENCABEZADO_X,  # CONSTANTE: Posición X fija
                               doc.height + doc.topMargin - 15,  # CONSTANTE: Posición Y fija
                               width=self.ENCABEZADO_WIDTH,    # CONSTANTE: Ancho fijo
                               height=self.ENCABEZADO_HEIGHT,  # CONSTANTE: Alto fijo
                               preserveAspectRatio=True)
            else:
                self._encabezado_texto(canvas, doc)
            
            # ==========================================
            # PIE DE PÁGINA (parte inferior) - COORDENADAS FIJAS
            # ==========================================
            if os.path.exists(self.ruta_pie):
                canvas.drawImage(self.ruta_pie, 
                               self.PIE_X,      # CONSTANTE: Posición X fija
                               self.PIE_Y,      # CONSTANTE: Posición Y fija
                               width=self.PIE_WIDTH,   # CONSTANTE: Ancho fijo
                               height=self.PIE_HEIGHT, # CONSTANTE: Alto fijo
                               preserveAspectRatio=True)
            else:
                self._pie_texto(canvas, doc)
                
        except Exception as e:
            print(f"Error en encabezado primera página: {e}")
        finally:
            canvas.restoreState()
    
    def paginas_siguientes(self, canvas, doc):
        """Encabezado y pie para páginas siguientes - EXACTAMENTE IGUALES a primera página"""
        canvas.saveState()
        try:
            # ==========================================
            # ENCABEZADO (parte superior) - MISMAS COORDENADAS QUE PRIMERA PÁGINA
            # ==========================================
            if os.path.exists(self.ruta_imagen):
                canvas.drawImage(self.ruta_imagen, 
                               self.ENCABEZADO_X,  # MISMA posición X que primera página
                               doc.height + doc.topMargin - 15,  # MISMA posición Y que primera página
                               width=self.ENCABEZADO_WIDTH,    # MISMO ancho que primera página
                               height=self.ENCABEZADO_HEIGHT,  # MISMA altura que primera página
                               preserveAspectRatio=True)
            else:
                self._encabezado_texto(canvas, doc)
            
            # ==========================================
            # PIE DE PÁGINA (parte inferior) - MISMAS COORDENADAS QUE PRIMERA PÁGINA
            # ==========================================
            if os.path.exists(self.ruta_pie):
                canvas.drawImage(self.ruta_pie, 
                               self.PIE_X,      # MISMA posición X que primera página
                               self.PIE_Y,      # MISMA posición Y que primera página
                               width=self.PIE_WIDTH,   # MISMO ancho que primera página
                               height=self.PIE_HEIGHT, # MISMA altura que primera página
                               preserveAspectRatio=True)
            else:
                self._pie_texto(canvas, doc)
                
        except Exception as e:
            print(f"Error en encabezado páginas siguientes: {e}")
        finally:
            canvas.restoreState()
    
    def _encabezado_texto(self, canvas, doc):
        """Encabezado de texto como fallback - MISMA posición siempre"""
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredText(doc.width/2 + 72, doc.height + doc.topMargin - 30, 
                              "GOBIERNO AUTÓNOMO MUNICIPAL DE TUPIZA")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredText(doc.width/2 + 72, doc.height + doc.topMargin - 45,
                              "Planta Municipal de Premoldeados")
        canvas.drawCentredText(doc.width/2 + 72, doc.height + doc.topMargin - 60,
                              "Sistema de Control de Inventarios")
    
    def _pie_texto(self, canvas, doc):
        """Pie de página de texto como fallback - MISMA posición siempre"""
        canvas.setFont("Helvetica", 8)
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        canvas.drawCentredText(doc.width/2 + 72, 40,
                              f"Generado el {fecha_actual}")
        canvas.drawCentredText(doc.width/2 + 72, 25,
                              "Sistema de Control de Inventarios - Planta Premoldeados Tupiza")

class MarcaDeAgua(Flowable):
    """Clase para crear marca de agua en PDFs - CORREGIDA para evitar conflictos con márgenes"""
    
    def __init__(self, texto="GOBIERNO MUNICIPAL TUPIZA"):
        if not PDF_DISPONIBLE:
            return
        Flowable.__init__(self)
        self.texto = texto
        # CORRIGIDO: Usar dimensiones más pequeñas para evitar conflicto con márgenes
        self.width = 400  # Reducido de letter[0] (612)
        self.height = 300  # Reducido de letter[1] (792)
    
    def draw(self):
        """Dibuja la marca de agua - CORREGIDA para evitar errores de tamaño"""
        if not PDF_DISPONIBLE:
            return
        canvas = self.canv
        canvas.saveState()
        try:
            canvas.setFont("Helvetica", 50)  # Reducido de 60
            canvas.setFillGray(0.95)  # Más claro (era 0.9)
            canvas.rotate(45)
            # CORRIGIDO: Posición ajustada para no interferir con márgenes
            canvas.drawCentredText(200, -50, self.texto)  # Posición más conservadora
        except Exception as e:
            print(f"Error dibujando marca de agua: {e}")
        finally:
            canvas.restoreState()

class PDFCreator:
    """
    Generador avanzado de reportes PDF con encabezado y pie de página institucional
    """
    
    @staticmethod
    def crear_estilos():
        """Crea estilos personalizados para el PDF"""
        if not PDF_DISPONIBLE:
            return None
            
        estilos = getSampleStyleSheet()
        
        # Estilo para título principal
        titulo_principal = ParagraphStyle(
            'TituloPrincipal',
            parent=estilos['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4e79'),
            alignment=TA_CENTER,
            spaceAfter=20,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para subtítulos
        subtitulo = ParagraphStyle(
            'Subtitulo',
            parent=estilos['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2e75b6'),
            spaceBefore=15,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para encabezado de sección
        encabezado = ParagraphStyle(
            'Encabezado',
            parent=estilos['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#1f4e79'),
            alignment=TA_CENTER,
            spaceBefore=5,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        # Estilo para pie de página
        pie = ParagraphStyle(
            'Pie',
            parent=estilos['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceBefore=10,
            fontName='Helvetica'
        )
        
        return {
            'titulo': titulo_principal,
            'subtitulo': subtitulo,
            'encabezado': encabezado,
            'pie': pie,
            'normal': estilos['Normal']
        }
    
    @staticmethod
    def generar_pdf_materiales(con_encabezado=True, con_marca_agua=False):
        """Genera reporte PDF completo de materiales
        
        Args:
            con_encabezado (bool): Si True, incluye encabezado y pie institucional en TODAS las páginas
            con_marca_agua (bool): Si True, incluye marca de agua (puede causar problemas con márgenes)
        """
        if not PDF_DISPONIBLE:
            print("❌ ReportLab no disponible")
            print("💡 Instala con: pip install reportlab")
            return None
        
        try:
            # Verificar que existe el archivo
            if not os.path.exists(ARCHIVO_EXCEL_MATERIALES):
                print(f"❌ No se encontró el archivo: {ARCHIVO_EXCEL_MATERIALES}")
                return None
            
            # Nombre del archivo PDF
            nombre_pdf = f"reporte_materiales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Configurar encabezado según parámetro
            if con_encabezado:
                # Crear el documento con márgenes ajustados para encabezado y pie
                doc = SimpleDocTemplate(
                    nombre_pdf,
                    pagesize=letter,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=120,  # Espacio para encabezado
                    bottomMargin=100  # Espacio para pie de página
                )
                
                # Crear instancia del encabezado personalizado
                encabezado_personalizado = EncabezadoPersonalizado()
            else:
                # Documento simple sin encabezado
                doc = SimpleDocTemplate(
                    nombre_pdf,
                    pagesize=letter,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=72,   # Márgenes normales
                    bottomMargin=72 # Márgenes normales
                )
            
            elementos = []
            estilos = PDFCreator.crear_estilos()
            
            # Agregar marca de agua solo si se solicita específicamente
            if con_encabezado and con_marca_agua:
                try:
                    elementos.append(MarcaDeAgua())
                except Exception as e:
                    print(f"⚠️ Error agregando marca de agua (se omite): {e}")
            
            # Título del documento
            titulo_doc = Paragraph("REPORTE DETALLADO DE INVENTARIO DE MATERIALES", estilos['titulo'])
            elementos.append(titulo_doc)
            elementos.append(Spacer(1, 20))
            
            # Información del reporte
            info_reporte = f"""
            <b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            <b>Sistema:</b> Control de Inventarios - Versión Modular<br/>
            <b>Entidad:</b> Gobierno Autónomo Municipal de Tupiza<br/>
            <b>Departamento:</b> Planta Municipal de Premoldeados
            """
            elementos.append(Paragraph(info_reporte, estilos['normal']))
            elementos.append(Spacer(1, 30))
            
            # Resumen ejecutivo
            elementos.append(Paragraph("1. RESUMEN EJECUTIVO", estilos['subtitulo']))
            
            # Obtener estadísticas
            total_registros = ExcelManager.contar_registros_materiales()
            stock_actual = ExcelManager.obtener_stock_materiales()
            
            resumen = f"""
            <b>Estado del Inventario:</b><br/>
            • Total de registros de movimientos: {total_registros}<br/>
            • Número de materiales diferentes: {len(stock_actual)}<br/>
            • Fecha de última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            • Status del sistema: Operativo
            """
            elementos.append(Paragraph(resumen, estilos['normal']))
            elementos.append(Spacer(1, 20))
            
            # Stock actual
            elementos.append(Paragraph("2. STOCK ACTUAL DE MATERIALES", estilos['subtitulo']))
            
            if stock_actual:
                # Crear tabla de stock
                datos_stock = [['Material', 'Cantidad Actual', 'Estado']]
                
                for material, cantidad in stock_actual.items():
                    if cantidad < 10:
                        estado = "🔴 Crítico"
                    elif cantidad < 50:
                        estado = "🟡 Bajo"
                    else:
                        estado = "🟢 Normal"
                    
                    datos_stock.append([material, f"{cantidad:.2f}", estado])
                
                tabla_stock = Table(datos_stock, colWidths=[3*inch, 1.5*inch, 1.5*inch])
                tabla_stock.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elementos.append(tabla_stock)
            else:
                elementos.append(Paragraph("No hay datos de stock disponibles.", estilos['normal']))
            
            elementos.append(Spacer(1, 30))
            
            # Últimos movimientos
            elementos.append(Paragraph("3. ÚLTIMOS MOVIMIENTOS", estilos['subtitulo']))
            
            try:
                import openpyxl
                libro = openpyxl.load_workbook(ARCHIVO_EXCEL_MATERIALES)
                hoja = libro.active
                
                datos_movimientos = [['Fecha', 'Material', 'Tipo', 'Cantidad', 'Observaciones']]
                
                # Obtener últimos 10 registros
                max_row = hoja.max_row
                start_row = max(5, max_row - 9)  # Últimos 10 registros
                
                for row in range(start_row, max_row + 1):
                    fecha = hoja.cell(row=row, column=1).value or ""
                    material = hoja.cell(row=row, column=3).value or ""
                    tipo = hoja.cell(row=row, column=5).value or ""
                    cantidad = hoja.cell(row=row, column=6).value or 0
                    observaciones = hoja.cell(row=row, column=7).value or ""
                    
                    datos_movimientos.append([
                        str(fecha), str(material), str(tipo), 
                        f"{cantidad:.2f}", str(observaciones)[:30] + "..." if len(str(observaciones)) > 30 else str(observaciones)
                    ])
                
                libro.close()
                
                tabla_movimientos = Table(datos_movimientos, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 2*inch])
                tabla_movimientos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e75b6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elementos.append(tabla_movimientos)
                
            except Exception as e:
                elementos.append(Paragraph(f"Error al cargar movimientos: {e}", estilos['normal']))
            
            elementos.append(Spacer(1, 30))
            
            # Estadísticas generales
            elementos.append(Paragraph("4. ESTADÍSTICAS GENERALES", estilos['subtitulo']))
            
            estadisticas = f"""
            <b>INFORMACIÓN DEL SISTEMA:</b><br/>
            • Total de registros en el sistema: {total_registros}<br/>
            • Archivo de datos: {os.path.basename(ARCHIVO_EXCEL_MATERIALES)}<br/>
            • Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            • Estado del sistema: Operativo<br/>
            <br/>
            <b>RECOMENDACIONES:</b><br/>
            • Realizar backup semanal de los archivos<br/>
            • Verificar niveles críticos de materiales<br/>
            • Mantener registro actualizado de movimientos<br/>
            • Revisar consumos anómalos<br/>
            <br/>
            <b>CONTACTO TÉCNICO:</b><br/>
            • Sistema: Control de Inventarios Modular<br/>
            • Soporte: Secretaría Municipal Técnica<br/>
            • Ubicación: Planta Municipal de Premoldeados - Tupiza
            """
            
            elementos.append(Paragraph(estadisticas, estilos['normal']))
            
            # Construir PDF con o sin encabezado según configuración
            if con_encabezado:
                doc.build(elementos, 
                         onFirstPage=encabezado_personalizado.primera_pagina,     # ✅ Encabezado en primera página
                         onLaterPages=encabezado_personalizado.paginas_siguientes) # ✅ Encabezado en TODAS las páginas siguientes
                print(f"✅ PDF CON ENCABEZADO generado exitosamente: {nombre_pdf}")
            else:
                doc.build(elementos)  # PDF simple sin encabezado
                print(f"✅ PDF SIMPLE generado exitosamente: {nombre_pdf}")
            
            return nombre_pdf
            
        except Exception as e:
            print(f"❌ Error generando PDF de materiales: {e}")
            return None
    
    @staticmethod
    def generar_pdf_combustibles(con_encabezado=True, con_marca_agua=False):
        """Genera reporte específico de combustibles
        
        Args:
            con_encabezado (bool): Si True, incluye encabezado y pie institucional en TODAS las páginas
            con_marca_agua (bool): Si True, incluye marca de agua (puede causar problemas con márgenes)
        """
        if not PDF_DISPONIBLE:
            print("❌ ReportLab no disponible")
            print("💡 Instala con: pip install reportlab")
            return None
        
        try:
            nombre_pdf = f"reporte_combustibles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Configurar encabezado según parámetro
            if con_encabezado:
                # Crear el documento con márgenes ajustados
                doc = SimpleDocTemplate(
                    nombre_pdf,
                    pagesize=letter,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=120,  # Espacio para encabezado
                    bottomMargin=100  # Espacio para pie de página
                )
                
                # Crear instancia del encabezado personalizado
                encabezado_personalizado = EncabezadoPersonalizado()
            else:
                # Documento simple sin encabezado
                doc = SimpleDocTemplate(
                    nombre_pdf,
                    pagesize=letter,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=72,   # Márgenes normales
                    bottomMargin=72 # Márgenes normales
                )
            
            elementos = []
            estilos = PDFCreator.crear_estilos()
            
            # Agregar marca de agua solo si se solicita específicamente
            if con_encabezado and con_marca_agua:
                try:
                    elementos.append(MarcaDeAgua())
                except Exception as e:
                    print(f"⚠️ Error agregando marca de agua (se omite): {e}")
            
            # Título
            titulo = Paragraph("REPORTE ESPECIALIZADO DE COMBUSTIBLES", estilos['titulo'])
            elementos.append(titulo)
            elementos.append(Spacer(1, 20))
            
            # Información del reporte
            info_reporte = f"""
            <b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
            <b>Tipo de reporte:</b> Control de Combustibles<br/>
            <b>Sistema:</b> Inventario Modular - Tupiza<br/>
            <b>Departamento:</b> Planta Municipal de Premoldeados
            """
            elementos.append(Paragraph(info_reporte, estilos['normal']))
            elementos.append(Spacer(1, 30))
            
            # Obtener datos de combustibles - CORREGIDO: método correcto
            stock_combustibles = ExcelManager.obtener_datos_combustibles()  # ✅ Método correcto
            
            if stock_combustibles:
                elementos.append(Paragraph("STOCK ACTUAL DE COMBUSTIBLES", estilos['subtitulo']))
                
                # Crear tabla de combustibles
                datos_combustibles = [['Tipo de Combustible', 'Cantidad (Litros)', 'Estado', 'Nivel']]
                
                for combustible, cantidad in stock_combustibles.items():
                    if cantidad < 50:
                        estado = "🔴 Crítico"
                        nivel = "Requiere abastecimiento inmediato"
                    elif cantidad < 100:
                        estado = "🟡 Bajo"
                        nivel = "Programar abastecimiento"
                    else:
                        estado = "🟢 Normal"
                        nivel = "Stock adecuado"
                    
                    datos_combustibles.append([
                        combustible.title(),
                        f"{cantidad:.2f}",
                        estado,
                        nivel
                    ])
                
                tabla_combustibles = Table(datos_combustibles, colWidths=[1.5*inch, 1.2*inch, 1*inch, 2.3*inch])
                tabla_combustibles.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elementos.append(tabla_combustibles)
                elementos.append(Spacer(1, 30))
                
                # Análisis
                total_combustible = sum(stock_combustibles.values())
                elementos.append(Paragraph("ANÁLISIS DEL STOCK", estilos['subtitulo']))
                
                analisis = f"""
                <b>Total de combustible en stock:</b> {total_combustible:.2f} litros<br/>
                <b>Gasolina:</b> {stock_combustibles.get('gasolina', 0):.2f} litros<br/>
                <b>Diesel:</b> {stock_combustibles.get('diesel', 0):.2f} litros<br/>
                <br/>
                <b>Estado general:</b> {'🔴 Crítico' if total_combustible < 100 else '🟢 Aceptable'}<br/>
                <b>Recomendación:</b> {'Abastecimiento urgente requerido' if total_combustible < 100 else 'Mantener monitoreo regular'}
                """
                
                elementos.append(Paragraph(analisis, estilos['normal']))
            else:
                elementos.append(Paragraph("No hay datos de combustibles disponibles.", estilos['normal']))
            
            # Construir PDF con o sin encabezado según configuración
            if con_encabezado:
                doc.build(elementos,
                         onFirstPage=encabezado_personalizado.primera_pagina,     # ✅ Encabezado en primera página
                         onLaterPages=encabezado_personalizado.paginas_siguientes) # ✅ Encabezado en TODAS las páginas siguientes
                print(f"✅ PDF DE COMBUSTIBLES CON ENCABEZADO generado: {nombre_pdf}")
            else:
                doc.build(elementos)  # PDF simple sin encabezado
                print(f"✅ PDF DE COMBUSTIBLES SIMPLE generado: {nombre_pdf}")
            
            return nombre_pdf
            
        except Exception as e:
            print(f"❌ Error generando PDF de combustibles: {e}")
            return None

# ============================================================================
# FUNCIONES DE UTILIDAD PARA PDFS
# ============================================================================

def validar_reportlab():
    """Valida que ReportLab esté disponible"""
    return PDF_DISPONIBLE

def crear_imagenes_ejemplo():
    """Crea imágenes de ejemplo para encabezado y pie si no existen"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        # Crear imagen de encabezado
        if not os.path.exists("encabezado_tupiza.png"):
            img_encabezado = Image.new('RGB', (700, 120), color='white')
            draw = ImageDraw.Draw(img_encabezado)
            
            # Título principal
            draw.text((50, 20), "GOBIERNO AUTÓNOMO MUNICIPAL DE TUPIZA", fill='blue')
            draw.text((50, 50), "Planta Municipal de Premoldeados", fill='black')
            draw.text((50, 80), "Sistema de Control de Inventarios", fill='gray')
            
            img_encabezado.save("encabezado_tupiza.png")
            print("✅ Imagen de encabezado creada: encabezado_tupiza.png")
        
        # Crear imagen de pie
        if not os.path.exists("pie_tupiza.png"):
            img_pie = Image.new('RGB', (700, 80), color='lightgray')
            draw = ImageDraw.Draw(img_pie)
            
            fecha = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            draw.text((50, 20), f"Generado el {fecha}", fill='black')
            draw.text((50, 50), "Documento oficial del Gobierno Municipal de Tupiza", fill='blue')
            
            img_pie.save("pie_tupiza.png")
            print("✅ Imagen de pie creada: pie_tupiza.png")
        
    except ImportError:
        print("⚠️ PIL no disponible para crear imágenes de ejemplo")
        print("💡 Las imágenes se crearán como texto si no existen")

# Verificación inicial
if __name__ == "__main__":
    print("🔍 Verificando módulo PDF...")
    if validar_reportlab():
        print("✅ ReportLab disponible")
        crear_imagenes_ejemplo()
        
        print("\n📋 EJEMPLOS DE USO:")
        print("# Reportes CON encabezado y pie institucional (RECOMENDADO)")
        print("PDFCreator.generar_pdf_materiales()  # ✅ Con encabezado")
        print("PDFCreator.generar_pdf_combustibles()  # ✅ Con encabezado")
        print("")
        print("# Reportes CON encabezado Y marca de agua (OPCIONAL)")
        print("PDFCreator.generar_pdf_materiales(con_marca_agua=True)  # Con marca de agua")
        print("PDFCreator.generar_pdf_combustibles(con_marca_agua=True)  # Con marca de agua")
        print("")
        print("# Reportes SIN encabezado (si es necesario)")
        print("PDFCreator.generar_pdf_materiales(con_encabezado=False)  # Sin encabezado")
        print("PDFCreator.generar_pdf_combustibles(con_encabezado=False)  # Sin encabezado")
        print("")
        print("🎯 SOLUCIONES IMPLEMENTADAS:")
        print("✅ Encabezado y pie en TODAS las páginas (no solo la primera)")
        print("✅ MISMO tamaño exacto en primera página y páginas siguientes")
        print("✅ Coordenadas fijas mediante constantes para evitar variaciones")
        print("✅ Opción configurable para incluir o no encabezado")
        print("✅ Marca de agua opcional y segura (sin conflictos de tamaño)")
        print("✅ Método correcto para datos de combustibles")
        print("✅ Fallback a texto si las imágenes no existen")
        
    else:
        print("❌ ReportLab no disponible")
        print("💡 Instala con: pip install reportlab")
