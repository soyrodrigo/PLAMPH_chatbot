#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 DIAGNÓSTICO ESPECÍFICO DE CEMENTO
===================================

Este script verificará por qué la gráfica de cemento dice que no hay datos
suficientes cuando sí hay registros de cemento en el Excel.
"""

import os
import openpyxl
from datetime import datetime

def buscar_archivo_excel():
    """Busca el archivo Excel en las ubicaciones posibles"""
    ubicaciones = [
        "datos/inventario_materiales.xlsx",
        "inventario_materiales.xlsx",
        "./datos/inventario_materiales.xlsx"
    ]
    
    for ubicacion in ubicaciones:
        if os.path.exists(ubicacion):
            return ubicacion
    
    return None

def analizar_datos_cemento_detallado():
    """Analiza específicamente los datos de cemento como lo hace la función original"""
    
    archivo = buscar_archivo_excel()
    if not archivo:
        print("❌ No se encontró archivo Excel")
        return
    
    print(f"📁 Analizando archivo: {archivo}")
    
    try:
        libro = openpyxl.load_workbook(archivo)
        hoja = libro.active
        
        print(f"📊 Total de filas: {hoja.max_row}")
        print(f"📊 Total de columnas: {hoja.max_column}")
        
        # Mostrar encabezados para verificar estructura
        print(f"\n📋 ESTRUCTURA DEL ARCHIVO:")
        for col in range(1, min(hoja.max_column + 1, 8)):
            valor = hoja.cell(row=4, column=col).value
            print(f"   Columna {col}: {valor}")
        
        print(f"\n🔍 BUSCANDO DATOS DE CEMENTO...")
        print("La función busca:")
        print("• Material que contenga 'Cemento'")
        print("• Movimiento que contenga 'Salida'") 
        print("• Cantidad numérica válida")
        
        datos_cemento_encontrados = []
        todos_los_datos_cemento = []
        
        # Revisar TODOS los registros de cemento (no solo salidas)
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value  # Columna C
            movimiento = hoja.cell(row=row, column=5).value  # Columna E
            cantidad = hoja.cell(row=row, column=6).value   # Columna F
            fecha = hoja.cell(row=row, column=1).value      # Columna A
            
            # Si hay material y contiene "Cemento"
            if material and "Cemento" in str(material):
                todos_los_datos_cemento.append({
                    'fila': row,
                    'fecha': fecha,
                    'material': material,
                    'movimiento': movimiento,
                    'cantidad': cantidad
                })
                
                print(f"   📦 Fila {row}: {material} | {movimiento} | {cantidad} | {fecha}")
                
                # Verificar si cumple criterios específicos para gráfica de cemento
                if movimiento and "Salida" in str(movimiento):
                    try:
                        cantidad_num = float(str(cantidad).replace(",", "."))
                        datos_cemento_encontrados.append((fecha, cantidad_num))
                        print(f"      ✅ VÁLIDO para gráfica: {fecha} - {cantidad_num}")
                    except Exception as e:
                        print(f"      ❌ Error procesando cantidad '{cantidad}': {e}")
                else:
                    print(f"      ⚠️ No es salida: {movimiento}")
        
        print(f"\n📊 RESUMEN:")
        print(f"• Total registros de cemento encontrados: {len(todos_los_datos_cemento)}")
        print(f"• Registros válidos para gráfica (salidas): {len(datos_cemento_encontrados)}")
        
        if len(todos_los_datos_cemento) == 0:
            print("❌ PROBLEMA: No se encontró ningún registro de cemento")
            print("💡 Verifica que el material se llame exactamente 'Cemento'")
        elif len(datos_cemento_encontrados) == 0:
            print("❌ PROBLEMA: Se encontraron registros de cemento pero ninguna SALIDA")
            print("💡 La gráfica de cemento solo muestra SALIDAS (consumo)")
            print("💡 Necesitas registrar algunas salidas de cemento")
        else:
            print("✅ Datos válidos encontrados para la gráfica")
            
        return datos_cemento_encontrados, todos_los_datos_cemento
        
    except Exception as e:
        print(f"❌ Error leyendo archivo: {e}")
        return None, None

def simular_funcion_grafica_cemento():
    """Simula exactamente lo que hace la función original de gráfica de cemento"""
    
    print(f"\n🎬 === SIMULANDO FUNCIÓN ORIGINAL ===")
    
    archivo = buscar_archivo_excel()
    if not archivo:
        print("❌ No se encontró archivo - función retornaría None")
        return None
    
    try:
        libro = openpyxl.load_workbook(archivo)
        hoja = libro.active
        
        datos_cemento = []
        
        # Esta es la lógica EXACTA de la función original
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value
            if material and "Cemento" in str(material):
                fecha = hoja.cell(row=row, column=1).value
                movimiento = hoja.cell(row=row, column=5).value
                cantidad = hoja.cell(row=row, column=6).value
                
                try:
                    cantidad_num = float(str(cantidad).replace(",", "."))
                    if "Salida" in str(movimiento):
                        datos_cemento.append((fecha, cantidad_num))
                        print(f"✅ Datos válidos: {fecha} - {cantidad_num}")
                except:
                    print(f"⚠️ Error procesando: {material} | {movimiento} | {cantidad}")
                    continue
        
        print(f"\n📊 RESULTADO DE LA SIMULACIÓN:")
        print(f"• Datos de cemento encontrados: {len(datos_cemento)}")
        
        if not datos_cemento:
            print("❌ La función retornaría None (no hay datos)")
            print("💡 Esto explica el mensaje 'no hay datos suficientes'")
            return None
        else:
            print("✅ La función generaría la gráfica")
            print("🔍 Datos que usaría:")
            for fecha, cantidad in datos_cemento:
                print(f"   📈 {fecha}: {cantidad} bolsas")
            return datos_cemento
        
    except Exception as e:
        print(f"❌ Error en simulación: {e}")
        return None

def verificar_con_graphics_generator():
    """Verifica qué está haciendo realmente el GraphicsGenerator"""
    
    print(f"\n🔧 === VERIFICANDO GRAPHICS_GENERATOR ===")
    
    try:
        from modules.graphics_generator import GraphicsGenerator
        
        print("✅ GraphicsGenerator importado correctamente")
        
        # Intentar generar la gráfica y capturar el resultado
        print("🎯 Intentando generar gráfica de cemento...")
        
        resultado = GraphicsGenerator.generar_grafica_cemento()
        
        if resultado:
            print(f"✅ Gráfica generada: {resultado}")
            if os.path.exists(resultado):
                tamaño = os.path.getsize(resultado) / 1024
                print(f"   📏 Tamaño: {tamaño:.1f} KB")
                print("💡 La gráfica se generó correctamente")
            else:
                print("⚠️ Se devolvió un nombre pero el archivo no existe")
        else:
            print("❌ GraphicsGenerator retornó None")
            print("💡 Esto confirma que no encuentra datos válidos")
            
        return resultado
        
    except Exception as e:
        print(f"❌ Error con GraphicsGenerator: {e}")
        import traceback
        traceback.print_exc()
        return None

def recomendar_solucion(datos_cemento_validos, todos_datos_cemento):
    """Recomienda la solución según el problema encontrado"""
    
    print(f"\n💡 === RECOMENDACIONES ===")
    
    if not todos_datos_cemento:
        print("🎯 PROBLEMA: No hay registros de cemento")
        print("✅ SOLUCIÓN:")
        print("   1. Usa el bot para registrar material 'Cemento'")
        print("   2. Asegúrate de escribir exactamente 'Cemento'")
        print("   3. Registra algunas entradas y salidas")
        
    elif not datos_cemento_validos:
        print("🎯 PROBLEMA: Tienes cemento pero no hay SALIDAS")
        print("✅ SOLUCIÓN:")
        print("   1. La gráfica de cemento muestra CONSUMO (salidas)")
        print("   2. Registra algunas salidas de cemento:")
        print("      📦 Registrar Material → Cemento → 📉 Salida → [cantidad]")
        print("   3. Ejemplos de salidas:")
        print("      • Producción de adoquines: 25 bolsas")
        print("      • Construcción: 10 bolsas")
        print("      • Reparaciones: 5 bolsas")
        
    else:
        print("🎯 PROBLEMA: Datos válidos existen pero la función no los ve")
        print("✅ SOLUCIÓN:")
        print("   1. Puede ser problema de configuración de rutas")
        print("   2. O problema en el módulo GraphicsGenerator")
        print("   3. Verifica que el archivo esté en la ubicación correcta")

def main():
    """Función principal del diagnóstico"""
    print("🔍 DIAGNÓSTICO ESPECÍFICO DE GRÁFICA DE CEMENTO")
    print("=" * 60)
    print(f"⏰ Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # 1. Análisis detallado
    datos_validos, todos_datos = analizar_datos_cemento_detallado()
    
    # 2. Simulación de función original
    datos_simulacion = simular_funcion_grafica_cemento()
    
    # 3. Verificación con módulo real
    resultado_real = verificar_con_graphics_generator()
    
    # 4. Recomendaciones
    recomendar_solucion(datos_validos, todos_datos)
    
    # 5. Resumen final
    print(f"\n📋 === RESUMEN FINAL ===")
    if todos_datos:
        print(f"📦 Registros de cemento: {len(todos_datos)}")
        if datos_validos:
            print(f"📈 Salidas válidas: {len(datos_validos)}")
            print("🎯 ESPERADO: La gráfica debería funcionar")
        else:
            print("📈 Salidas válidas: 0")
            print("🎯 ESPERADO: Mensaje 'no hay datos suficientes'")
    else:
        print("📦 No hay registros de cemento")
        print("🎯 ESPERADO: Mensaje 'no hay datos suficientes'")
    
    print(f"⏰ Diagnóstico completado: {datetime.now().strftime('%H:%M:%S')}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠️ Diagnóstico interrumpido")
    except Exception as e:
        print(f"\n❌ Error en diagnóstico: {e}")
        import traceback
        traceback.print_exc()
