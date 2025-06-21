#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📊 GRÁFICA DE CEMENTO CORREGIDA - DATOS REALES
=============================================

Este script encontrará TUS datos reales de cemento y generará
la gráfica específica de consumo diario de cemento en bolsas.
"""

import os
import openpyxl
from datetime import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from collections import defaultdict

def buscar_archivo_excel():
    """Busca el archivo Excel en las ubicaciones posibles"""
    ubicaciones = [
        "datos/inventario_materiales.xlsx",
        "inventario_materiales.xlsx",
        "./datos/inventario_materiales.xlsx"
    ]
    
    for ubicacion in ubicaciones:
        if os.path.exists(ubicacion):
            return ubicacion
    
    return None

def mostrar_todos_los_datos():
    """Muestra TODOS los datos del Excel para identificar los de cemento"""
    
    archivo = buscar_archivo_excel()
    if not archivo:
        print("❌ No se encontró archivo Excel")
        return None
    
    print(f"📁 Analizando archivo: {archivo}")
    
    try:
        libro = openpyxl.load_workbook(archivo)
        hoja = libro.active
        
        print(f"📊 Total de filas: {hoja.max_row}")
        
        # Mostrar encabezados
        print(f"\n📋 ENCABEZADOS (Fila 4):")
        encabezados = []
        for col in range(1, hoja.max_column + 1):
            valor = hoja.cell(row=4, column=col).value
            encabezados.append(valor)
            print(f"   Columna {col}: {valor}")
        
        print(f"\n📋 TODOS LOS DATOS DEL EXCEL:")
        print("=" * 80)
        
        datos_encontrados = []
        
        for row in range(5, hoja.max_row + 1):
            fila_datos = []
            for col in range(1, min(hoja.max_column + 1, 8)):  # Primeras 7 columnas
                valor = hoja.cell(row=row, column=col).value
                fila_datos.append(valor)
            
            # Si hay algún dato en la fila
            if any(fila_datos):
                datos_encontrados.append((row, fila_datos))
                
                # Mostrar la fila
                fecha = fila_datos[0] if len(fila_datos) > 0 else ""
                hora = fila_datos[1] if len(fila_datos) > 1 else ""
                material = fila_datos[2] if len(fila_datos) > 2 else ""
                usuario = fila_datos[3] if len(fila_datos) > 3 else ""
                movimiento = fila_datos[4] if len(fila_datos) > 4 else ""
                cantidad = fila_datos[5] if len(fila_datos) > 5 else ""
                
                print(f"Fila {row:2d}: {fecha} | {hora} | {material} | {usuario} | {movimiento} | {cantidad}")
                
                # Verificar si contiene cemento
                if material and "cemento" in str(material).lower():
                    print(f"   🎯 ¡CEMENTO ENCONTRADO! → {material}")
        
        print("=" * 80)
        print(f"📊 Total de filas con datos: {len(datos_encontrados)}")
        
        return datos_encontrados
        
    except Exception as e:
        print(f"❌ Error leyendo archivo: {e}")
        return None

def buscar_cemento_flexible():
    """Busca datos de cemento con mayor flexibilidad"""
    
    archivo = buscar_archivo_excel()
    if not archivo:
        return []
    
    print(f"\n🔍 === BÚSQUEDA FLEXIBLE DE CEMENTO ===")
    
    try:
        libro = openpyxl.load_workbook(archivo)
        hoja = libro.active
        
        datos_cemento_encontrados = []
        
        # Buscar en todas las filas
        for row in range(5, hoja.max_row + 1):
            material = hoja.cell(row=row, column=3).value  # Columna C
            movimiento = hoja.cell(row=row, column=5).value  # Columna E
            cantidad = hoja.cell(row=row, column=6).value   # Columna F
            fecha = hoja.cell(row=row, column=1).value      # Columna A
            
            # Búsqueda flexible de cemento
            if material:
                material_lower = str(material).lower()
                
                # Diferentes variantes de cemento
                variantes_cemento = ['cemento', 'cement', 'cemto', 'simento']
                
                if any(variante in material_lower for variante in variantes_cemento):
                    print(f"   📦 Cemento encontrado en fila {row}: {material}")
                    print(f"      📅 Fecha: {fecha}")
                    print(f"      📊 Movimiento: {movimiento}")
                    print(f"      🔢 Cantidad: {cantidad}")
                    
                    # Si es una salida (consumo)
                    if movimiento:
                        movimiento_lower = str(movimiento).lower()
                        
                        # Diferentes formas de indicar salida
                        indicadores_salida = ['salida', 'consumo', 'uso', 'gasto', 'produccion', 'producción', '📉']
                        
                        if any(indicador in movimiento_lower for indicador in indicadores_salida):
                            try:
                                # Procesar cantidad
                                cantidad_str = str(cantidad).replace(",", ".").strip()
                                cantidad_num = float(cantidad_str)
                                
                                # Procesar fecha
                                if isinstance(fecha, datetime):
                                    fecha_str = fecha.strftime("%d/%m/%Y")
                                else:
                                    fecha_str = str(fecha) if fecha else "Sin fecha"
                                
                                datos_cemento_encontrados.append({
                                    'fecha': fecha_str,
                                    'fecha_obj': fecha,
                                    'cantidad': cantidad_num,
                                    'material': material,
                                    'movimiento': movimiento,
                                    'fila': row
                                })
                                
                                print(f"      ✅ VÁLIDO: {fecha_str} - {cantidad_num} bolsas")
                                
                            except Exception as e:
                                print(f"      ❌ Error procesando cantidad '{cantidad}': {e}")
                        else:
                            print(f"      ⚠️ No es salida: {movimiento}")
                    else:
                        print(f"      ⚠️ Sin movimiento definido")
        
        print(f"\n📊 RESULTADO BÚSQUEDA FLEXIBLE:")
        print(f"• Datos de cemento (salidas) encontrados: {len(datos_cemento_encontrados)}")
        
        return datos_cemento_encontrados
        
    except Exception as e:
        print(f"❌ Error en búsqueda flexible: {e}")
        return []

def generar_grafica_cemento_personalizada(datos_cemento):
    """Genera la gráfica específica de consumo de cemento por día"""
    
    if not datos_cemento:
        print("❌ No hay datos de cemento para generar gráfica")
        return None
    
    print(f"\n📊 === GENERANDO GRÁFICA DE CEMENTO ===")
    print(f"Datos a graficar: {len(datos_cemento)} registros")
    
    try:
        # Agrupar por fecha
        consumo_por_fecha = defaultdict(float)
        
        for dato in datos_cemento:
            fecha = dato['fecha']
            cantidad = dato['cantidad']
            consumo_por_fecha[fecha] += cantidad
            print(f"   📅 {fecha}: +{cantidad} bolsas")
        
        # Preparar datos para la gráfica
        fechas = list(consumo_por_fecha.keys())
        cantidades = list(consumo_por_fecha.values())
        
        print(f"\n📈 DATOS AGRUPADOS POR FECHA:")
        for fecha, total in consumo_por_fecha.items():
            print(f"   📅 {fecha}: {total} bolsas")
        
        # Crear gráfica
        plt.figure(figsize=(12, 8))
        
        # Configurar colores y estilo
        plt.style.use('default')
        color_principal = '#2E86AB'
        color_fondo = '#F8F9FA'
        
        # Crear gráfica de barras
        bars = plt.bar(fechas, cantidades, color=color_principal, alpha=0.8, width=0.6)
        
        # Personalizar gráfica
        plt.title('📊 CONSUMO DIARIO DE CEMENTO\n🏭 Planta Municipal de Premoldeados - Tupiza', 
                 fontsize=16, fontweight='bold', pad=20)
        plt.xlabel('Fecha', fontsize=14, fontweight='bold')
        plt.ylabel('Cemento Consumido (bolsas)', fontsize=14, fontweight='bold')
        
        # Rotar etiquetas de fechas
        plt.xticks(rotation=45, ha='right')
        
        # Agregar valores encima de las barras
        for bar, cantidad in zip(bars, cantidades):
            altura = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., altura + max(cantidades)*0.01,
                    f'{cantidad:.0f}', ha='center', va='bottom', fontweight='bold', fontsize=11)
        
        # Grid para mejor lectura
        plt.grid(True, alpha=0.3, axis='y', linestyle='--')
        
        # Información adicional
        total_consumo = sum(cantidades)
        promedio_diario = total_consumo / len(cantidades) if cantidades else 0
        
        plt.figtext(0.02, 0.02, 
                   f'Total consumido: {total_consumo:.0f} bolsas | Promedio diario: {promedio_diario:.1f} bolsas', 
                   fontsize=10, style='italic')
        
        # Ajustar diseño
        plt.tight_layout()
        
        # Guardar gráfica
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"consumo_cemento_{timestamp}.png"
        
        plt.savefig(nombre_archivo, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        print(f"✅ Gráfica generada: {nombre_archivo}")
        
        if os.path.exists(nombre_archivo):
            tamaño = os.path.getsize(nombre_archivo) / 1024
            print(f"📏 Tamaño del archivo: {tamaño:.1f} KB")
            
        return nombre_archivo
        
    except Exception as e:
        print(f"❌ Error generando gráfica: {e}")
        import traceback
        traceback.print_exc()
        return None

def corregir_modulo_graphics_generator():
    """Sugiere cómo corregir el módulo GraphicsGenerator"""
    
    print(f"\n🔧 === CORRECCIÓN DEL MÓDULO ===")
    print("Para que el bot funcione correctamente, necesitas actualizar")
    print("el archivo modules/graphics_generator.py con una búsqueda más flexible.")
    print("\nEl problema actual:")
    print("• Busca exactamente 'Cemento' (sensible a mayúsculas)")
    print("• Solo busca en columnas específicas")
    print("• No considera variantes del nombre")
    print("\nSolución:")
    print("• Búsqueda insensible a mayúsculas/minúsculas")
    print("• Múltiples variantes de 'cemento'")
    print("• Búsqueda más flexible de 'salida'")

def main():
    """Función principal"""
    print("📊 GRÁFICA DE CEMENTO - VERSIÓN CORREGIDA")
    print("=" * 60)
    print(f"⏰ Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print()
    print("🎯 OBJETIVO: Generar gráfica de consumo diario de cemento")
    print("📊 Mostrará: Cemento usado en producción por día (bolsas)")
    
    # 1. Mostrar todos los datos del Excel
    print(f"\n📋 PASO 1: Revisar todos los datos del Excel")
    todos_datos = mostrar_todos_los_datos()
    
    if not todos_datos:
        print("❌ No se pudieron leer los datos del Excel")
        return
    
    # 2. Búsqueda flexible de cemento
    print(f"\n🔍 PASO 2: Búsqueda flexible de datos de cemento")
    datos_cemento = buscar_cemento_flexible()
    
    if not datos_cemento:
        print("❌ No se encontraron datos de cemento")
        print("💡 Verifica que:")
        print("   • El material contenga la palabra 'cemento'")
        print("   • El movimiento indique salida/consumo/uso")
        print("   • La cantidad sea un número válido")
        return
    
    # 3. Generar gráfica personalizada
    print(f"\n📊 PASO 3: Generar gráfica personalizada")
    archivo_grafica = generar_grafica_cemento_personalizada(datos_cemento)
    
    if archivo_grafica:
        print(f"\n🎉 === ÉXITO ===")
        print(f"✅ Gráfica generada: {archivo_grafica}")
        print(f"📊 Datos procesados: {len(datos_cemento)} registros")
        print(f"📈 Consumo total: {sum(d['cantidad'] for d in datos_cemento):.0f} bolsas")
        print(f"💡 Abre el archivo para ver la gráfica")
    else:
        print(f"\n❌ === ERROR ===")
        print("No se pudo generar la gráfica")
    
    # 4. Sugerencias para corregir el bot
    corregir_modulo_graphics_generator()
    
    print(f"\n⏰ Finalizado: {datetime.now().strftime('%H:%M:%S')}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
